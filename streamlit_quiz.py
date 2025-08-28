# streamlit_quiz.py
import streamlit as st
import os
import json
import random
from collections import defaultdict
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# ==============================
# 공통 유틸
# ==============================
def reset_to_home():
    st.session_state.app_mode = 'setup'
    st.session_state.show_answer = False
    st.session_state.submitted = False
    st.session_state.step = 0
    st.session_state.score = 0
    st.session_state.finished = False
    st.session_state.last_judged = None

# 사이드바 공통 메뉴
st.sidebar.markdown("## 🚀 메뉴")
if st.sidebar.button("🏠 처음으로", key="btn_home_sidebar"):
    reset_to_home()
    st.rerun()

# 네비게이션: 퀴즈 / 표 변환
nav = st.sidebar.radio(
    "기능 선택",
    options=["퀴즈", "표 변환(JSON→Excel)"],
    index=0,
    horizontal=False,
    key="nav_mode",
)

# ==============================
# 표 변환 기능 (표_변환.py 통합)
# ==============================
# 유형 매핑
ref_map = {
    "table_ref": "표 설명 문장",
    "row_ref": "행 설명 문장",
    "col_ref": "열 설명 문장",
    "cell_ref": "불연속 영역 설명 문장"
}

def extract_mdfcn_values(obj, sep="\n"):
    """
    mdfcn_infos가 문자열/리스트/딕셔너리 등 어떤 형태든
    내부 'value' 값만 추출해 중복 제거(등장 순서 유지) 후 sep로 연결.
    'mdfcn_memo'가 JSON 문자열이면 파싱해서 'value'만 추출.
    'table_ref', 'row_ref', 'col_ref', 'cell_ref' 등 type 태그 문자열은 무시.
    """
    values = []
    TYPE_TAGS = {"table_ref", "row_ref", "col_ref", "cell_ref"}

    def _dedup_keep_order(items):
        seen = set()
        out = []
        for it in items:
            if it and it not in seen:
                seen.add(it)
                out.append(it)
        return out

    def _walk(x):
        if x is None:
            return

        if isinstance(x, str):
            s = x.strip()
            if not s:
                return
            if s[:1] in ("[", "{"):
                try:
                    _walk(json.loads(s))
                    return
                except Exception:
                    if s not in TYPE_TAGS:
                        values.append(s)
                    return
            if s not in TYPE_TAGS:
                values.append(s)
            return

        if isinstance(x, dict):
            v = x.get("value")
            if isinstance(v, str):
                v = v.strip()
                if v:
                    values.append(v)

            mm = x.get("mdfcn_memo")
            if isinstance(mm, str):
                mm_s = mm.strip()
                if mm_s:
                    try:
                        _walk(json.loads(mm_s))
                    except Exception:
                        pass

            for k, sub in x.items():
                if k in ("value", "mdfcn_memo"):
                    continue
                if isinstance(sub, (list, dict)):
                    _walk(sub)
            return

        if isinstance(x, (list, tuple)):
            for it in x:
                _walk(it)
            return

    _walk(obj)
    return sep.join(_dedup_keep_order([v for v in values if isinstance(v, str)]))

def _extract_url_from_metadata(meta):
    if isinstance(meta, dict):
        u = meta.get("url", "")
        if isinstance(u, list):
            return u[0] if u else ""
        return str(u or "")
    return ""

def convert_json_to_excel_bytes(data):
    """
    입력 JSON(dict)을 표 형식으로 변환하여 서식 포함 XLSX bytes와
    DataFrame(미리보기 용)을 반환.
    """
    rows = []
    group_counts = defaultdict(int)

    docs = data.get("document", [])
    for doc in docs:
        doc_id = doc.get("id", "")
        worker = str(doc.get("worker_id_cnst") or "").strip()
        metadata = doc.get("metadata", {})

        mdfcn_raw = doc.get("mdfcn_infos", doc.get("mdfcn_memo", []))
        mdfcn_text = extract_mdfcn_values(mdfcn_raw, sep="\n")

        url = _extract_url_from_metadata(metadata)
        ex_list = doc.get("EX", [])

        for ex in ex_list:
            ref_type = ex.get("reference", {}).get("reference_type", "")
            exp_list = ex.get("exp_sentence", [])

            for exp in exp_list:
                sentence = ""
                try:
                    sent_list = exp.get("설명 문장", [])
                    sentence = sent_list[0] if sent_list else ""
                except Exception:
                    sentence = ""

                rows.append({
                    "id": doc_id,
                    "worker_id_cnst": worker,
                    "유형": ref_map.get(ref_type, ref_type),
                    "설명 문장": sentence,
                    "metadata": json.dumps(metadata, ensure_ascii=False, indent=2),
                    "mdfcn_infos": mdfcn_text,
                    "url": url,
                })
                group_counts[doc_id] += 1

    # 빈 문서 방어
    if not rows:
        df_empty = pd.DataFrame(columns=["id", "worker_id_cnst", "유형", "설명 문장", "metadata", "mdfcn_infos"])
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_empty.to_excel(writer, index=False, sheet_name="sheet1")
        buf.seek(0)
        return buf.getvalue(), df_empty

    df = pd.DataFrame(
        rows,
        columns=["id", "worker_id_cnst", "유형", "설명 문장", "metadata", "mdfcn_infos", "url"]
    ).fillna("")

    # url 컬럼은 엑셀 표에는 숨기고, 하이퍼링크 지정에만 사용
    df_out = df.drop(columns=["url"])

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="sheet1")
        ws = writer.sheets["sheet1"]

        # 열 너비
        widths = {"A": 18, "B": 16, "C": 13, "D": 80, "E": 50, "F": 50}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # 헤더 스타일
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # metadata, mdfcn_infos 헤더 강조
        ws["E1"].fill = PatternFill("solid", fgColor="BDD7EE")
        ws["F1"].fill = PatternFill("solid", fgColor="FCE4D6")

        # 데이터 영역 서식
        thin = Side(style="thin", color="999999")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        max_row = ws.max_row
        max_col = ws.max_column

        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if c >= 4:  # D,E,F
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(vertical="top")
                cell.border = border

        # 병합: A(id), B(worker_id_cnst), E(metadata), F(mdfcn_infos)
        cur_row = 2
        # id별 행 개수 재계산 (rows 순서 기준)
        id_order_counts = defaultdict(int)
        for r in rows:
            id_order_counts[r["id"]] += 1

        for doc_id, count in id_order_counts.items():
            if count > 1:
                for col in (1, 2, 5, 6):
                    ws.merge_cells(
                        start_row=cur_row, start_column=col,
                        end_row=cur_row + count - 1, end_column=col
                    )
                    top_cell = ws.cell(row=cur_row, column=col)
                    top_cell.alignment = Alignment(vertical="top", wrap_text=True)
                    for rr in range(cur_row, cur_row + count):
                        ws.cell(row=rr, column=col).border = border
            cur_row += count

        # 하이퍼링크: id 블록 첫 행의 E열에만 부여
        first_row_for_id = {}
        for idx, row in enumerate(rows, start=2):
            first_row_for_id.setdefault(row["id"], idx)

        for idx, row in enumerate(rows, start=2):
            url = row.get("url", "")
            if not url:
                continue
            if idx != first_row_for_id[row["id"]]:
                continue
            meta_cell = ws.cell(row=idx, column=5)
            meta_cell.hyperlink = url

        # 행 높이 자동 보정
        for r in range(2, max_row + 1):
            max_lines = 1
            for c in (4, 5, 6):  # D,E,F
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and "\n" in val:
                    lines = val.count("\n") + 1
                    if lines > max_lines:
                        max_lines = lines
            ws.row_dimensions[r].height = min(15 + (max_lines - 1) * 12, 200)

        # 틀 고정
        ws.freeze_panes = "A2"

    buf.seek(0)
    # 미리보기용으로는 주요 컬럼만
    preview_cols = ["id", "worker_id_cnst", "유형", "설명 문장", "mdfcn_infos"]
    preview_df = df_out[preview_cols].copy()

    return buf.getvalue(), preview_df

# ==============================
# 표 변환 화면
# ==============================
if nav == "표 변환(JSON→Excel)":
    st.markdown("<h1 style='color:#0066CC'>📄 표 변환: JSON → Excel</h1>", unsafe_allow_html=True)
    st.markdown("JSON 파일 경로를 직접 입력하거나, 아래 업로더로 파일을 올린 뒤 변환을 실행하세요.")

    json_path = st.text_input("JSON 경로 직접 입력", value=st.session_state.get("json_path", ""), placeholder="/path/to/project_141.json")
    st.session_state.json_path = json_path

    uploaded = st.file_uploader("또는 JSON 파일 업로드", type=["json"])

    run = st.button("🔁 변환 실행")
    if run:
        data = None
        # 업로드 우선
        if uploaded is not None:
            try:
                data = json.loads(uploaded.getvalue().decode("utf-8"))
            except Exception as e:
                st.error(f"업로드된 JSON 파싱 오류: {e}")
        elif json_path.strip():
            try:
                with open(json_path.strip(), "r", encoding="utf-8") as f:
                    data = json.load(f)
            except Exception as e:
                st.error(f"경로에서 JSON 로드 실패: {e}")
        else:
            st.warning("JSON 경로를 입력하거나, 파일을 업로드하세요.")

        if data is not None:
            try:
                xlsx_bytes, preview_df = convert_json_to_excel_bytes(data)
                st.success("변환 완료")
                st.dataframe(preview_df.head(20), use_container_width=True)
                st.download_button(
                    label="📥 Excel 다운로드 (.xlsx)",
                    data=xlsx_bytes,
                    file_name="표_변환.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"변환 중 오류: {e}")

    st.stop()

# ==============================
# 퀴즈 화면
# ==============================
# 상태 기본값
if 'app_mode' not in st.session_state:
    st.session_state.app_mode = 'setup'
if 'selected_subject' not in st.session_state:
    st.session_state.selected_subject = None
if 'order_mode' not in st.session_state:
    st.session_state.order_mode = "랜덤"
if 'solve_mode' not in st.session_state:
    st.session_state.solve_mode = "한 문제씩(즉시 채점)"
if 'last_judged' not in st.session_state:
    st.session_state.last_judged = None

# 데이터 폴더 및 파일 리스트 (퀴즈 전용)
DATA_DIR = './quiz_data'
files = [f for f in os.listdir(DATA_DIR) if f.endswith('.json')] if os.path.isdir(DATA_DIR) else []
subjects = [os.path.splitext(f)[0] for f in files]
if not subjects:
    st.error("❌ quiz_data 폴더에 json 파일이 없습니다.")
    st.stop()
if st.session_state.selected_subject is None:
    st.session_state.selected_subject = subjects[0]

# 1. 첫 화면: 옵션 선택
if st.session_state.app_mode == 'setup':
    st.markdown("<h1 style='color:#0066CC'>📝 영양교육 객관식 퀴즈</h1>", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("#### 1. 과목(파일) 선택")
    subject = st.selectbox("과목(파일)", subjects, index=subjects.index(st.session_state.selected_subject))
    st.session_state.selected_subject = subject

    st.markdown("#### 2. 문제 순서")
    order_mode = st.radio("문제 순서", ["랜덤", "순차"], index=0 if st.session_state.order_mode == "랜덤" else 1, horizontal=True)
    st.session_state.order_mode = order_mode

    st.markdown("#### 3. 풀이 모드")
    solve_mode = st.radio("풀이 모드", ["한 문제씩(즉시 채점)", "모의고사(최종 제출)"], index=0 if st.session_state.solve_mode.startswith("한 문제씩") else 1, horizontal=True)
    st.session_state.solve_mode = solve_mode

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🚩 문제풀이 시작", use_container_width=True):
        filepath = os.path.join(DATA_DIR, subject + '.json')
        with open(filepath, encoding='utf-8') as f:
            questions = json.load(f)
        st.session_state.questions = questions
        indices = list(range(len(questions)))
        if order_mode == "랜덤":
            random.shuffle(indices)
        st.session_state.quiz_order = indices
        st.session_state.step = 0
        st.session_state.score = 0
        st.session_state.last_input = ""
        st.session_state.show_answer = False
        st.session_state.inputs = [None] * len(questions)
        st.session_state.answered = [False] * len(questions)
        st.session_state.finished = False
        st.session_state.submitted = False
        st.session_state.last_judged = None
        st.session_state.app_mode = 'quiz'
        st.rerun()
    st.stop()

# 2. 퀴즈풀이 화면
questions = st.session_state.questions
order = st.session_state.quiz_order
step = st.session_state.step
score = st.session_state.score
solve_mode = st.session_state.solve_mode
inputs = st.session_state.inputs

if st.session_state.app_mode == 'quiz':
    st.markdown(f"<h2 style='color:#0066CC'>[{st.session_state.selected_subject}] 퀴즈풀이</h2>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown(f"<div style='padding:8px 0 0 0; color:#333'><b>문제 {step+1} / {len(questions)}</b></div>", unsafe_allow_html=True)
    st.progress((step + 1) / len(questions) if step < len(questions) else 1.0)
    st.markdown(f"<span style='color: #16a34a; font-weight:700;'>현재 점수: {score}</span>", unsafe_allow_html=True)

# 한 문제씩(즉시 채점)
if solve_mode == "한 문제씩(즉시 채점)" and not st.session_state.finished:
    if step >= len(order):
        st.session_state.finished = True
        st.rerun()

    idx = order[step]
    q = questions[idx]
    st.markdown("-----")
    st.markdown(f"<b style='font-size:1.1em;'>{q['question']}</b>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    index_val = q["choices"].index(inputs[idx]) if inputs[idx] in q["choices"] else None
    choice = st.radio("정답을 선택하세요", q["choices"], key=f"choice_{step}", index=index_val)
    inputs[idx] = choice

    answer_btn_col, submit_col = st.columns([1, 1])
    with answer_btn_col:
        pass
    with submit_col:
        if st.button("✅ 제출", key=f"submit_{step}", use_container_width=True):
            if not choice:
                st.warning("정답을 선택해 주세요!")
            else:
                if choice == q["answer"]:
                    st.success("🎉 정답입니다! 대단해요!")
                    st.balloons()
                    st.session_state.last_judged = 'correct'
                    st.session_state.score += 1
                else:
                    st.error("😥 오답입니다... 조금만 더 힘내요!")
                    st.snow()
                    st.session_state.last_judged = 'wrong'
                st.session_state.answered[idx] = True
                st.session_state.submitted = True
                st.session_state.show_answer = True
                st.rerun()

    if st.session_state.show_answer:
        st.info(f"정답: {q['answer']}")

    if st.session_state.get('last_judged') == 'correct':
        st.markdown("<div style='font-size:2em;'>🎊 🎉 🥳</div>", unsafe_allow_html=True)
    elif st.session_state.get('last_judged') == 'wrong':
        st.markdown("<div style='font-size:2em;'>😭 ❄️ 💧</div>", unsafe_allow_html=True)

    if st.session_state.get('submitted', False):
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("➡️ 다음 문제", use_container_width=True):
            st.session_state.step += 1
            st.session_state.submitted = False
            st.session_state.last_judged = None
            st.session_state.show_answer = False
            st.rerun()

    if step >= len(questions) - 1 and st.session_state.answered[idx]:
        st.session_state.finished = True
        st.rerun()

# 모의고사(최종 제출)
elif solve_mode == "모의고사(최종 제출)" and not st.session_state.finished:
    st.markdown("-----")
    st.write("모든 문제에 답을 입력한 뒤 최종 제출 버튼을 누르세요.")
    for i, idx in enumerate(order):
        q = questions[idx]
        with st.expander(f"문제 {i+1}: {q['question']}"):
            index_val = q["choices"].index(inputs[idx]) if inputs[idx] in q["choices"] else None
            choice = st.radio("정답을 선택하세요", q["choices"], key=f"mock_choice_{i}", index=index_val)
            inputs[idx] = choice

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("📊 최종 제출", use_container_width=True):
        score = 0
        for i, idx in enumerate(order):
            user_choice = inputs[idx]
            q = questions[idx]
            if user_choice == q["answer"]:
                score += 1
        st.session_state.score = score
        st.session_state.finished = True
        st.rerun()

# 3. 결과
if st.session_state.finished:
    st.balloons()
    st.markdown(f"<h2 style='color:#0066CC;'>🥳 퀴즈 완료!</h2>", unsafe_allow_html=True)
    st.markdown(f"<div style='font-size:1.5em; color:#333;'>총 <span style='color:#2563eb;font-weight:700;'>{len(questions)}</span>문제 중 <span style='color:#16a34a;font-weight:700;'>{st.session_state.score}</span>개 맞추셨습니다!</div>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button("🔄 다시 시작", use_container_width=True):
            reset_to_home()
            st.rerun()
    with col2:
        if st.button("🏠 처음으로", key="btn_home_final", use_container_width=True):
            reset_to_home()
            st.rerun()
