#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
JSON -> Excel 변환기 (문서 단위 병합 + URL 하이퍼링크)
- 입력: 단일 JSON 파일 또는 폴더(하위 모든 *.json 재귀 탐색)
- 열: id / worker_id_cnst / Medium_category / 유형 / 설명 문장 / metadata / mdfcn_memo(검수자 수정 이력)
- 같은 document.id 묶음에서 [id, worker_id_cnst, Medium_category, metadata, mdfcn_memo] 세로 병합
- metadata: JSON 스타일 멀티라인(plain 텍스트). 셀 자체에 하이퍼링크(파란색, 밑줄 없음)
- mdfcn_memo: 문자열 JSON 파싱 후 "작업 목록 1, 2, 3…" 재번호
"""

import json
import math
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ===== 경로 하드코딩 =====
INPUT_PATH = Path('/Users/dataly/Desktop/사진 변환 1차(250827)/project_146.json')   # 폴더 또는 단일 JSON 파일
OUTPUT_XLSX = Path("output.xlsx")
# =======================

META_ORDER = [
    "note", "image", "copyright", "term_id", "Major_category",
    "title", "url", "Medium_category", "domain", "media_id",
    "publisher", "term", "source_id",
]

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")
LINK_BLUE = "0563C1"

# ---------- 파일 유틸 ----------
def iter_json_files(path: Path) -> Iterable[Path]:
    if path.is_file() and path.suffix.lower() == ".json":
        yield path
    elif path.is_dir():
        yield from path.rglob("*.json")

def load_json(path: Path) -> Optional[Dict[str, Any]]:
    for enc in ("utf-8", "utf-8-sig"):
        try:
            return json.loads(path.read_text(encoding=enc))
        except Exception:
            pass
    print(f"❌ JSON 로드 실패: {path}")
    return None

# ---------- 파싱 ----------
TYPE_BRACKET_RE = re.compile(r"^\s*\[(.+?)\]\s*(.*)$")

def extract_sentences(doc: Dict[str, Any]) -> List[Tuple[str, str]]:
    out: List[Tuple[str, str]] = []
    for ex in doc.get("EX", []):
        for item in ex.get("exp_sentence", []) or []:
            if not isinstance(item, dict):
                continue
            for _k, v in item.items():
                seq = v if isinstance(v, list) else [v]
                for s in seq:
                    if not s:
                        continue
                    text = str(s).strip()
                    m = TYPE_BRACKET_RE.match(text)
                    if m:
                        out.append((m.group(1).strip(), m.group(2).strip()))
                    else:
                        out.append(("", text))
    return out

def _clean_url(u: str) -> str:
    if not u:
        return ""
    u = str(u).strip()
    if (u.startswith('"') and u.endswith('"')) or (u.startswith("'") and u.endswith("'")):
        u = u[1:-1].strip()
    return u

def format_metadata_and_url(meta: Dict[str, Any]) -> Tuple[str, str]:
    """메타데이터 멀티라인 문자열과 url만 분리 반환"""
    url_only = _clean_url(meta.get("url", ""))
    lines = ['metadata : {']
    for k in META_ORDER:
        v = meta.get(k, "")
        if k == "url":
            v = url_only or meta.get("url", "")
        lines.append(f'  "{k}": "{v}",' if k != META_ORDER[-1] else f'  "{k}": "{v}"')
    lines.append("}")
    return "\n".join(lines), url_only

def extract_mdfcn_memo(mdfcn_infos):
    if not mdfcn_infos:
        return ""
    out, idx = [], 1
    for info in mdfcn_infos:
        raw = info.get("mdfcn_memo", "")
        if not raw:
            continue
        try:
            arr = json.loads(raw)
            if isinstance(arr, list):
                for obj in arr:
                    val = str((obj or {}).get("value", "")).strip()
                    if val:
                        out.append(f"작업 목록 {idx} : {val}")
                        idx += 1
        except Exception:
            val = str(raw).strip()
            if val:
                out.append(f"작업 목록 {idx} : {val}")
                idx += 1
    # ✅ 항목 사이에 빈 줄 추가
    return "\n\n".join(out)


def to_rows(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    docs = data.get("document", [])
    if not isinstance(docs, list):
        return rows

    for doc in docs:
        img_id = str(doc.get("id", ""))
        meta = doc.get("metadata", {}) or {}
        medium_category = str(meta.get("Medium_category", "") or "")
        worker_id_cnst = str(doc.get("worker_id_cnst", "") or "")

        md_text, md_url = format_metadata_and_url(meta)
        memo_text = extract_mdfcn_memo(doc.get("mdfcn_infos", []) or [])

        pairs = extract_sentences(doc) or [("", "")]
        for typ, sent in pairs:
            rows.append({
                "id": img_id,
                "worker_id_cnst": worker_id_cnst,
                "Medium_category": medium_category,
                "유형": typ,
                "설명 문장": sent,
                "metadata": md_text,
                "meta_url": md_url,  # 하이퍼링크용
                "mdfcn_memo(검수자 수정 이력)": memo_text,
            })
    return rows

# ---------- 엑셀 ----------
def estimate_wrapped_lines(text: str, col_chars: int) -> int:
    if not text:
        return 1
    total = 0
    width = max(col_chars, 5)
    for para in str(text).split("\n"):
        total += max(1, math.ceil(len(para) / (width * 1.08)))
    return max(1, total)

def write_excel(all_rows: List[Dict[str, Any]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "result"

    headers = [
        "id", "worker_id_cnst", "Medium_category",
        "유형", "설명 문장", "metadata", "mdfcn_memo\n(검수자 수정 이력)"
    ]
    ws.append(headers)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        cell.fill = HEADER_FILL

    widths = {1:12, 2:16, 3:14, 4:16, 5:80, 6:60, 7:50}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    start_row_by_group: Dict[Tuple[str], int] = {}
    count_by_group: Dict[Tuple[str], int] = {}

    current_row = 2
    for row in all_rows:
        ws.append([
            row.get("id",""),
            row.get("worker_id_cnst",""),
            row.get("Medium_category",""),
            row.get("유형",""),
            row.get("설명 문장",""),
            row.get("metadata",""),
            row.get("mdfcn_memo(검수자 수정 이력)",""),
        ])
        for c in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=c).alignment = Alignment(vertical="top", wrap_text=(c in (5,6,7)))
            ws.cell(row=current_row, column=c).border = THIN_BORDER

        key = (row.get("id",""),)
        if key not in start_row_by_group:
            start_row_by_group[key] = current_row
            count_by_group[key] = 0
        count_by_group[key] += 1
        current_row += 1

    # 병합
    merge_cols = [1,2,3,6,7]
    for key, start in start_row_by_group.items():
        cnt = count_by_group[key]
        if cnt > 1:
            end = start + cnt - 1
            for col in merge_cols:
                ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)
                ws.cell(row=start, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    # metadata 셀에 하이퍼링크(파란색, 밑줄 없음)
    # 같은 id 그룹의 "첫 행" 메타 셀만 처리
    first_url_by_id: Dict[str, str] = {}
    for r, row in enumerate(all_rows, start=2):
        rid = row.get("id","")
        if rid and rid not in first_url_by_id:
            first_url_by_id[rid] = row.get("meta_url","") or ""

    for key, start in start_row_by_group.items():
        doc_id = key[0]
        url = first_url_by_id.get(doc_id, "")
        if url and url.startswith(("http://", "https://")):
            c = ws.cell(row=start, column=6)
            c.hyperlink = url

            # 1) 파랑/밑줄 없이 (검정, 밑줄 없음)
            c.font = Font(color="000000", underline=None)

            # 혹시 Excel이 기본 'Hyperlink' 스타일을 강제로 입히면,
            # 아래 한 줄 추가 후 테두리/정렬 다시 적용하세요.
            # c.style = "Normal"
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = THIN_BORDER

    # 행 높이 자동
    LINE_HEIGHT_PT = 18
    group_starts = set(start_row_by_group.values())
    for r in range(2, current_row):
        desc = ws.cell(row=r, column=5).value or ""
        desc_lines = estimate_wrapped_lines(desc, widths[5])
        if r in group_starts:
            meta_plain = ws.cell(row=r, column=6).value or ""
            memo_plain = ws.cell(row=r, column=7).value or ""
            need = max(desc_lines,
                       estimate_wrapped_lines(meta_plain, widths[6]),
                       estimate_wrapped_lines(memo_plain, widths[7]))
        else:
            need = desc_lines
        ws.row_dimensions[r].height = max(1, need) * LINE_HEIGHT_PT

    wb.save(out_path)
    print(f"✅ 저장 완료: {out_path}")

# ---------- main ----------
def main():
    srcs = list(iter_json_files(INPUT_PATH))
    if not srcs:
        print("⚠️ 입력 경로에 JSON 파일이 없습니다.")
        return

    all_rows: List[Dict[str, Any]] = []
    for jf in srcs:
        data = load_json(jf)
        if data:
            all_rows.extend(to_rows(data))

    if not all_rows:
        print("⚠️ 변환할 행이 없습니다.")
        return

    write_excel(all_rows, OUTPUT_XLSX)

if __name__ == "__main__":
    main()
