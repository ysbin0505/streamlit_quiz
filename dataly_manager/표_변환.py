#표_변환.py
import json
import pandas as pd
from collections import defaultdict

from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# ===== 사용자 경로 =====
JSON_PATH = '/Users/data.ly/Desktop/일/08:27 (한림대 통합검수)/표_사진/표 변환 1차(250827)/project_141.json'
OUT_XLSX = '/Users/data.ly/Desktop/일/08:27 (한림대 통합검수)/표_사진/표 변환 1차(250827)/0827 표 및 사진 구축 말뭉치(표).xlsx'

# ===== 유형 매핑 =====
ref_map = {
    "table_ref": "표 설명 문장",
    "row_ref": "행 설명 문장",
    "col_ref": "열 설명 문장",
    "cell_ref": "불연속 영역 설명 문장"
}

# ===== mdfcn_infos 가공 헬퍼 =====
def extract_mdfcn_values(obj, sep="\n"):
    """
    mdfcn_infos가 문자열/리스트/딕셔너리 등 어떤 형태든
    내부의 'value' 값만 추출해 중복 제거(등장 순서 유지) 후 sep로 연결.
    - 'mdfcn_memo'가 JSON 문자열인 경우도 파싱해서 'value'만 추출.
    - 'table_ref', 'row_ref', 'col_ref', 'cell_ref' 등 type 태그 문자열은 무시.
    """
    values = []
    TYPE_TAGS = {"table_ref", "row_ref", "col_ref", "cell_ref"}

    def _dedup_keep_order(items):
        seen = set()
        out = []
        for it in items:
            if it and it not in seen:
                seen.add(it)
                out.append(it)
        return out

    def _walk(x):
        if x is None:
            return

        # 문자열 처리
        if isinstance(x, str):
            s = x.strip()
            if not s:
                return
            # JSON처럼 보이면 파싱 시도
            if s[:1] in ("[", "{"):
                try:
                    _walk(json.loads(s))
                    return
                except Exception:
                    # 파싱 실패 시: 태그 문자열이면 버리고, 아니면 값으로 취급
                    if s not in TYPE_TAGS:
                        values.append(s)
                    return
            # 일반 문자열: 태그면 무시, 아니면 값으로 수집
            if s not in TYPE_TAGS:
                values.append(s)
            return

        # dict 처리: 'value'만 수집, 'mdfcn_memo'만 파싱, 그 외 키는 list/dict일 때만 재귀
        if isinstance(x, dict):
            v = x.get("value")
            if isinstance(v, str):
                v = v.strip()
                if v:
                    values.append(v)

            mm = x.get("mdfcn_memo")
            if isinstance(mm, str):
                mm_s = mm.strip()
                if mm_s:
                    try:
                        _walk(json.loads(mm_s))
                    except Exception:
                        # mdfcn_memo가 JSON이 아니면 무시
                        pass

            # 나머지 키는 list/dict일 때에만 재귀 (문자열은 재귀 금지: 'type' 등 방지)
            for k, sub in x.items():
                if k in ("value", "mdfcn_memo"):
                    continue
                if isinstance(sub, (list, dict)):
                    _walk(sub)
            return

        # list/tuple 처리
        if isinstance(x, (list, tuple)):
            for it in x:
                _walk(it)
            return

    _walk(obj)
    return sep.join(_dedup_keep_order([v for v in values if isinstance(v, str)]))

# ===== JSON 로드 =====
with open(JSON_PATH, "r", encoding="utf-8") as f:
    data = json.load(f)

rows = []
group_counts = defaultdict(int)  # id별 행 개수 (엑셀 병합에 사용)

def extract_url(meta):
    # metadata가 dict일 때만 처리
    if isinstance(meta, dict):
        u = meta.get("url", "")
        if isinstance(u, list):
            return u[0] if u else ""
        return str(u or "")
    return ""

for doc in data.get("document", []):
    doc_id = doc.get("id", "")
    worker = str(doc.get("worker_id_cnst") or "").strip()
    metadata = doc.get("metadata", {})

    # mdfcn_infos 원본 수집 (키 변동 대비)
    mdfcn_raw = doc.get("mdfcn_infos", doc.get("mdfcn_memo", []))
    mdfcn_text = extract_mdfcn_values(mdfcn_raw, sep="\n")   # value당 개행

    for ex in doc.get("EX", []):
        ref_type = ex.get("reference", {}).get("reference_type", "")
        exp_list = ex.get("exp_sentence", [])

        for exp in exp_list:
            # 설명 문장: ["문장", ...] 형태일 수 있으므로 0번째 안전 추출
            sentence = ""
            try:
                sent_list = exp.get("설명 문장", [])
                sentence = sent_list[0] if sent_list else ""
            except Exception:
                sentence = ""

            rows.append({
                "id": doc_id,
                "worker_id_cnst": worker,
                "유형": ref_map.get(ref_type, ref_type),
                "설명 문장": sentence,
                "metadata": json.dumps(metadata, ensure_ascii=False, indent=2),
                "mdfcn_infos": mdfcn_text,  # 가공된 value 전용, 중복 제거, 개행
                "url": extract_url(metadata),   # 하이퍼링크용(엑셀에 컬럼으로 쓰진 않음)
            })
            group_counts[doc_id] += 1

# ===== DataFrame -> Excel 저장 =====
df = pd.DataFrame(
    rows,
    columns=["id", "worker_id_cnst", "유형", "설명 문장", "metadata", "mdfcn_infos"]  # url 컬럼은 제외
).fillna("")

with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="sheet1")
    ws = writer.sheets["sheet1"]

    # 열 너비
    widths = {
        "A": 18,  # id
        "B": 16,  # worker_id_cnst
        "C": 13,  # 유형
        "D": 80,  # 설명 문장
        "E": 50,  # metadata
        "F": 50,  # mdfcn_infos
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 머리글 스타일
    header_fill = PatternFill("solid", fgColor="D9E1F2")  # 연한 파랑
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # metadata, mdfcn_infos 헤더 강조
    ws["E1"].fill = PatternFill("solid", fgColor="BDD7EE")  # metadata
    ws["F1"].fill = PatternFill("solid", fgColor="FCE4D6")  # mdfcn_infos

    # 데이터 영역: 줄바꿈 + 상단 정렬 + 테두리
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if c >= 4:  # D, E, F
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
            cell.border = border

    # id, worker_id_cnst, metadata(E), mdfcn_infos(F) 병합 (같은 id 블록 기준)
    cur_row = 2
    for doc_id, count in group_counts.items():
        if count > 1:
            # A(id), B(worker_id_cnst), E(metadata), F(mdfcn_infos) 병합
            for col in (1, 2, 5, 6):
                ws.merge_cells(
                    start_row=cur_row, start_column=col,
                    end_row=cur_row + count - 1, end_column=col
                )
                # 병합된 셀 정렬/테두리
                top_cell = ws.cell(row=cur_row, column=col)
                top_cell.alignment = Alignment(vertical="top", wrap_text=True)
                for rr in range(cur_row, cur_row + count):
                    ws.cell(row=rr, column=col).border = border
        cur_row += count

    # 하이퍼링크: 병합 영역에서는 '맨 윗칸'에만 설정해야 오류가 나지 않음
    # id별 첫 행(row index)을 계산
    first_row_for_id = {}
    for idx, row in enumerate(rows, start=2):
        first_row_for_id.setdefault(row["id"], idx)

    # E열(metadata) 첫 행에만 링크 부여
    for idx, row in enumerate(rows, start=2):
        url = row.get("url", "")
        if not url:
            continue
        if idx != first_row_for_id[row["id"]]:
            continue  # 병합된 아래쪽 셀은 건너뜀 (MergedCell에 하이퍼링크 설정 불가)
        meta_cell = ws.cell(row=idx, column=5)  # E열
        meta_cell.hyperlink = url
        # meta_cell.style = "Hyperlink"  # 스타일은 적용하지 않음

    # 행 높이: D/E/F의 개행 수를 기준으로 근사 조절
    for r in range(2, max_row + 1):
        max_lines = 1
        for c in (4, 5, 6):  # D,E,F
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and "\n" in val:
                lines = val.count("\n") + 1
                if lines > max_lines:
                    max_lines = lines
        ws.row_dimensions[r].height = min(15 + (max_lines - 1) * 12, 200)

    # 틀 고정 (머리글 고정)
    ws.freeze_panes = "A2"

print(f"완료: {OUT_XLSX}")
