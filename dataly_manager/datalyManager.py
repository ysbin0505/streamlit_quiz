# datalyManager.py
import streamlit as st
import zipfile
import tempfile
import os
import json
import pandas as pd
from collections import defaultdict
from io import BytesIO

from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# 기존 기능용 모듈
from newspaper_eval_merged import json_to_excel_stacked
from newspaper_eval_json import merge_newspaper_eval

# ========== 공통 스타일 ==========
st.markdown("""
    <style>
    .main-title {font-size:2.1rem; font-weight:bold; color:#174B99; margin-bottom:0;}
    .sub-desc {font-size:1.1rem; color:#222;}
    .logo {height:60px; margin-bottom:15px;}
    .footer {color:#999; font-size:0.9rem; margin-top:40px;}
    div.stButton > button:first-child {background:#174B99; color:white; font-weight:bold; border-radius:8px;}
    .stTabs [data-baseweb="tab-list"] {background:#F6FAFD;}
    </style>
""", unsafe_allow_html=True)

# 상단 브랜드/로고
col1, col2 = st.columns([0.15, 0.85])
with col1:
    st.image("https://static.streamlit.io/examples/cat.jpg", width=55)
with col2:
    st.markdown('<div class="main-title">Dataly Manager</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-desc">업무 자동화, 평가 데이터 변환, 관리, 수합 웹앱</div>', unsafe_allow_html=True)

# ================== 표 변환(표_변환.py)에서 쓰는 헬퍼/매핑 ==================
ref_map = {
    "table_ref": "표 설명 문장",
    "row_ref": "행 설명 문장",
    "col_ref": "열 설명 문장",
    "cell_ref": "불연속 영역 설명 문장"
}

def extract_mdfcn_values(obj, sep="\n"):
    """
    mdfcn_infos가 문자열/리스트/딕셔너리 등 어떤 형태든
    내부의 'value' 값만 추출해 중복 제거(등장 순서 유지) 후 sep로 연결.
    - 'mdfcn_memo'가 JSON 문자열인 경우도 파싱해서 'value'만 추출.
    - 'table_ref', 'row_ref', 'col_ref', 'cell_ref' 등 type 태그 문자열은 무시.
    """
    values = []
    TYPE_TAGS = {"table_ref", "row_ref", "col_ref", "cell_ref"}

    def _dedup_keep_order(items):
        seen = set()
        out = []
        for it in items:
            if it and it not in seen:
                seen.add(it)
                out.append(it)
        return out

    def _walk(x):
        if x is None:
            return
        if isinstance(x, str):
            s = x.strip()
            if not s:
                return
            if s[:1] in ("[", "{"):
                try:
                    _walk(json.loads(s))
                    return
                except Exception:
                    if s not in TYPE_TAGS:
                        values.append(s)
                    return
            if s not in TYPE_TAGS:
                values.append(s)
            return
        if isinstance(x, dict):
            v = x.get("value")
            if isinstance(v, str):
                v = v.strip()
                if v:
                    values.append(v)
            mm = x.get("mdfcn_memo")
            if isinstance(mm, str):
                mm_s = mm.strip()
                if mm_s:
                    try:
                        _walk(json.loads(mm_s))
                    except Exception:
                        pass
            for k, sub in x.items():
                if k in ("value", "mdfcn_memo"):
                    continue
                if isinstance(sub, (list, dict)):
                    _walk(sub)
            return
        if isinstance(x, (list, tuple)):
            for it in x:
                _walk(it)
            return

    _walk(obj)
    return sep.join(_dedup_keep_order([v for v in values if isinstance(v, str)]))

def extract_url(meta):
    if isinstance(meta, dict):
        u = meta.get("url", "")
        if isinstance(u, list):
            return u[0] if u else ""
        return str(u or "")
    return ""

def convert_table_json_to_excel_bytes(data) -> bytes:
    """
    표_변환.py의 전체 로직을 메모리 버전으로 변환.
    입력: 파싱된 JSON(dict)
    출력: 생성된 XLSX의 bytes
    """
    rows = []
    group_counts = defaultdict(int)

    for doc in data.get("document", []):
        doc_id = doc.get("id", "")
        worker = str(doc.get("worker_id_cnst") or "").strip()
        metadata = doc.get("metadata", {})

        # mdfcn_infos 원본 수집 (키 변동 대비)
        mdfcn_raw = doc.get("mdfcn_infos", doc.get("mdfcn_memo", []))
        mdfcn_text = extract_mdfcn_values(mdfcn_raw, sep="\n")

        for ex in doc.get("EX", []):
            ref_type = ex.get("reference", {}).get("reference_type", "")
            exp_list = ex.get("exp_sentence", [])

            for exp in exp_list:
                sentence = ""
                try:
                    sent_list = exp.get("설명 문장", [])
                    sentence = sent_list[0] if sent_list else ""
                except Exception:
                    sentence = ""

                rows.append({
                    "id": doc_id,
                    "worker_id_cnst": worker,
                    "유형": ref_map.get(ref_type, ref_type),
                    "설명 문장": sentence,
                    "metadata": json.dumps(metadata, ensure_ascii=False, indent=2),
                    "mdfcn_infos": mdfcn_text,
                    "url": extract_url(metadata),
                })
                group_counts[doc_id] += 1

    df = pd.DataFrame(
        rows,
        columns=["id", "worker_id_cnst", "유형", "설명 문장", "metadata", "mdfcn_infos"]
    ).fillna("")

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="sheet1")
        ws = writer.sheets["sheet1"]

        # 열 너비
        widths = {"A": 18, "B": 16, "C": 13, "D": 80, "E": 50, "F": 50}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # 머리글 스타일
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # metadata, mdfcn_infos 헤더 강조
        ws["E1"].fill = PatternFill("solid", fgColor="BDD7EE")
        ws["F1"].fill = PatternFill("solid", fgColor="FCE4D6")

        thin = Side(style="thin", color="999999")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        max_row = ws.max_row
        max_col = ws.max_column

        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if c >= 4:  # D,E,F
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(vertical="top")
                cell.border = border

        # id 블록 병합: A, B, E, F
        cur_row = 2
        for doc_id, count in group_counts.items():
            if count > 1:
                for col in (1, 2, 5, 6):
                    ws.merge_cells(
                        start_row=cur_row, start_column=col,
                        end_row=cur_row + count - 1, end_column=col
                    )
                    top_cell = ws.cell(row=cur_row, column=col)
                    top_cell.alignment = Alignment(vertical="top", wrap_text=True)
                    for rr in range(cur_row, cur_row + count):
                        ws.cell(row=rr, column=col).border = border
            cur_row += count

        # 하이퍼링크: 병합 첫 행만
        first_row_for_id = {}
        for idx, row in enumerate(rows, start=2):
            first_row_for_id.setdefault(row["id"], idx)
        for idx, row in enumerate(rows, start=2):
            url = row.get("url", "")
            if not url:
                continue
            if idx != first_row_for_id[row["id"]]:
                continue
            meta_cell = ws.cell(row=idx, column=5)  # E
            meta_cell.hyperlink = url

        # 행 높이: D/E/F 개행 수 기준
        for r in range(2, max_row + 1):
            max_lines = 1
            for c in (4, 5, 6):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and "\n" in val:
                    lines = val.count("\n") + 1
                    if lines > max_lines:
                        max_lines = lines
            ws.row_dimensions[r].height = min(15 + (max_lines - 1) * 12, 200)

        # 틀 고정
        ws.freeze_panes = "A2"

    output.seek(0)
    return output.getvalue()

# ========== 탭 ==========
tabs = st.tabs([
    "🏠 홈",
    "📰 신문평가 수합",
    "💬 대화평가 병합",
    "📦 신문평가 병합",
    "📊 표 변환 (JSON→Excel)"  # 새 탭 추가
])

# 홈
with tabs[0]:
    st.markdown("#### 👋 환영합니다!<br>아래 탭에서 기능을 선택해 주세요.", unsafe_allow_html=True)
    st.markdown("""
    - 📰 신문평가 수합: 신문 JSON을 엑셀로 변환
    - 💬 대화평가 병합: 대화 평가 병합 (추가예정)
    - 📦 신문평가 병합: A/B팀 JSON 묶음 병합 ZIP 생성
    - 📊 표 변환: 단일 JSON을 표 형태 엑셀로 변환
    """)

# 신문평가 수합
with tabs[1]:
    st.header("📰 신문평가 JSON → 엑셀 자동 수합기")
    st.info("아래 순서대로 업로드 및 실행을 진행하세요.")
    uploaded_zip = st.file_uploader("1. 평가 데이터 ZIP 업로드 (폴더를 압축)", type=["zip"], key="file_upload_zip_sum")
    sum_week_num = st.number_input("2. 수합할 주차 (예: 1)", min_value=1, step=1, value=1, key="sum_week_num")
    storage_folder = st.selectbox("3. storage 폴더명 선택", ["storage0", "storage1"], key="sum_storage_folder")
    run_btn = st.button("실행 (엑셀 변환)", key="run_newspaper_sum")

    if uploaded_zip and run_btn:
        with tempfile.TemporaryDirectory() as temp_dir:
            tmp_zip_path = os.path.join(temp_dir, "data.zip")
            with open(tmp_zip_path, "wb") as f:
                f.write(uploaded_zip.read())
            with zipfile.ZipFile(tmp_zip_path, "r") as zip_ref:
                zip_ref.extractall(temp_dir)
            folder_list = [f for f in os.listdir(temp_dir) if os.path.isdir(os.path.join(temp_dir, f))]
            if not folder_list:
                st.error("압축파일 내부에 폴더가 없습니다. 폴더째 압축한 zip만 지원합니다.")
            else:
                root_path = os.path.join(temp_dir, folder_list[0])
                st.info("엑셀 변환을 시작합니다. (수초~수십초 소요)")
                json_to_excel_stacked(root_path, sum_week_num, storage_folder)
                excel_path = os.path.join(root_path, "summary_eval_all.xlsx")
                if os.path.exists(excel_path):
                    with open(excel_path, "rb") as f:
                        st.success("엑셀 변환 완료! 아래 버튼으로 다운로드하세요.")
                        st.download_button(
                            label="summary_eval_all.xlsx 다운로드",
                            data=f,
                            file_name="summary_eval_all.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                else:
                    st.error("엑셀 파일 생성 실패. 내부 오류를 확인하세요.")
    else:
        st.info("ZIP, 주차, 폴더명 입력 후 [실행]을 눌러주세요.")

# 대화평가 병합 (준비중)
with tabs[2]:
    st.header("💬 대화평가 병합 (준비중)")
    st.info("이 기능은 곧 추가됩니다. 원하시는 기능이 있다면 문의해 주세요.")

# 신문평가 병합
with tabs[3]:
    st.header("📦 신문평가 JSON 병합")
    st.info("ZIP 내 'A/A팀', 'B/B팀' 폴더와 JSON 파일이 존재해야 하며, 병합 결과는 자동으로 압축파일로 다운로드할 수 있습니다.")

    uploaded_zip = st.file_uploader("병합할 신문 원본 ZIP 업로드 (A/B팀 포함 폴더)", type=["zip"], key="merge_zip_upload")
    merge_week_num = st.number_input("병합할 주차 (예: 1)", min_value=1, step=1, value=1, key="merge_week_num")
    files_per_week = st.number_input("병합할 파일 수 (보통 102)", min_value=1, step=1, value=102, key="merge_files_per_week")
    run_merge_btn = st.button("신문평가 병합 실행", key="run_newspaper_merge")

    if uploaded_zip and run_merge_btn:
        with tempfile.TemporaryDirectory() as temp_dir:
            tmp_zip_path = os.path.join(temp_dir, "src.zip")
            with open(tmp_zip_path, "wb") as f:
                f.write(uploaded_zip.read())
            with zipfile.ZipFile(tmp_zip_path, "r") as zip_ref:
                zip_ref.extractall(temp_dir)

            st.write("압축 해제 후 임시폴더 목록:", os.listdir(temp_dir))
            candidate_dirs = [os.path.join(temp_dir, d) for d in os.listdir(temp_dir) if os.path.isdir(os.path.join(temp_dir, d))]
            if not candidate_dirs:
                st.error("압축 내부 폴더를 찾을 수 없습니다. ZIP 폴더 구조를 확인하세요.")
            else:
                base_dir = candidate_dirs[0]
                st.write("선택된 base_dir:", base_dir)
                st.write("base_dir 폴더 목록:", os.listdir(base_dir))

                with st.spinner("병합 중입니다..."):
                    msg, output_dir, out_zip_path = merge_newspaper_eval(
                        week_num=int(merge_week_num),
                        files_per_week=int(files_per_week),
                        base_dir=base_dir
                    )
                st.success(f"병합 결과: {msg}")
                with open(out_zip_path, "rb") as f:
                    st.download_button(
                        label=f"{merge_week_num}주차 병합 JSON ZIP 다운로드",
                        data=f,
                        file_name=f"merged_{merge_week_num}주차.zip",
                        mime="application/zip"
                    )
    else:
        st.info("ZIP 파일, 주차, 파일 수 입력 후 실행을 눌러주세요.")

# 📊 표 변환 (JSON→Excel) - 새 탭
with tabs[4]:
    st.header("📊 표 변환 (단일 JSON → Excel)")
    st.info("project_xxx.json 한 개를 업로드하면, 표 형태 엑셀로 변환하여 다운로드할 수 있습니다.")
    uploaded_json = st.file_uploader("JSON 업로드 (project_*.json)", type=["json"], key="table_json_upload")
    run_table_btn = st.button("엑셀 변환 실행", key="run_table_convert")

    if uploaded_json and run_table_btn:
        try:
            data = json.load(uploaded_json)
        except Exception as e:
            st.error(f"JSON 파싱 실패: {e}")
        else:
            if not isinstance(data, dict) or "document" not in data:
                st.error("JSON 최상위에 'document' 배열이 없습니다. 파일 구조를 확인하세요.")
            else:
                with st.spinner("엑셀 생성 중..."):
                    xlsx_bytes = convert_table_json_to_excel_bytes(data)
                st.success("엑셀 생성 완료! 아래 버튼으로 다운로드하세요.")
                st.download_button(
                    label="표_변환.xlsx 다운로드",
                    data=xlsx_bytes,
                    file_name="표_변환.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.info("JSON 파일 업로드 후 [엑셀 변환 실행]을 눌러주세요.")

# 하단 푸터
st.markdown("""
<hr>
<div class="footer">
문의: 검증 엔지니어 | Powered by Streamlit<br>
Copyright &copy; 2025. All rights reserved.
</div>
""", unsafe_allow_html=True)
