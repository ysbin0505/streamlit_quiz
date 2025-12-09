# -*- coding: utf-8 -*-
"""
사진(이미지) JSON -> Excel 변환기 (+ 엑셀 수정본을 JSON에 역반영)
- 입력: dict(JSON 파싱 결과)
- 출력: bytes(XLSX), datalyManager에서 st.download_button으로 바로 다운로드
- 같은 document.id 묶음에서 [id, worker_id_cnst, Medium_category, metadata, mdfcn_memo] 세로 병합
- metadata: 멀티라인 텍스트 + 같은 id 첫 행에만 URL 하이퍼링크(파란색, 밑줄 없음)

추가:
- apply_excel_desc_to_json_from_zip(zip_bytes, sheet_name=None, skip_blank=True)
  : ZIP(엑셀+단일 JSON)을 받아 엑셀의 '설명 문장'을 JSON에 반영해 반환
"""

import json
import math
import re
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple, Optional
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 표시 순서(메타 키)
META_ORDER = [
    "note", "image", "copyright", "term_id", "Major_category",
    "title", "url", "Medium_category", "domain", "media_id",
    "publisher", "term", "source_id",
]

_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

def xls_safe(val) -> str:
    """
    openpyxl이 허용하지 않는 XML 제어문자를 제거.
    숫자/None도 문자열로 안전 변환.
    """
    if val is None:
        return ""
    s = str(val)
    s = s.replace("\x00", "")
    s = _ILLEGAL_XML_RE.sub("", s)
    return s

META_NOTE_RE = re.compile(r'"note"\s*:\s*"(?P<note>.*?)"', re.DOTALL)

def _parse_metadata_cell(cell_text: Any) -> Dict[str, Any]:
    """
    'metadata : { ... }' 형태의 멀티라인 문자열에서 { ... } 만 추출하여 json.loads 시도.
    실패 시 최소한 "note"만 정규식으로 추출.
    """
    if cell_text is None:
        return {}
    s = str(cell_text).strip()
    if not s:
        return {}

    i, j = s.find("{"), s.rfind("}")
    if i == -1 or j == -1 or i >= j:
        s_fix = s.replace('""', '"')
        m = META_NOTE_RE.search(s_fix)
        return {"note": m.group("note")} if m else {}

    blob = s[i:j+1].strip()
    for candidate in (blob, blob.replace('""', '"')):
        try:
            return json.loads(candidate)
        except Exception:
            pass

    s_fix = blob.replace('""', '"')
    m = META_NOTE_RE.search(s_fix)
    return {"note": m.group("note")} if m else {}

def _collect_metadata_by_id(df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """
    엑셀 DF에서 id별로 metadata 셀을 파싱해 전체 metadata dict를 수집.
    - id는 ffill
    - 각 id에 대해 '비어있지 않은 첫 metadata dict'를 채택
    """
    if "id" not in df.columns or "metadata" not in df.columns:
        return {}

    tmp = df.copy()
    tmp["id"] = tmp["id"].ffill().astype(str)

    out: Dict[str, Dict[str, Any]] = {}
    for _, row in tmp.iterrows():
        _id = row["id"].strip()
        if not _id:
            continue
        meta_dict = _parse_metadata_cell(row["metadata"])
        if meta_dict and _id not in out:
            out[_id] = meta_dict
    return out

def _collect_note_by_id(df: pd.DataFrame) -> Dict[str, str]:
    """
    엑셀 DF에서 id별로 metadata 셀을 파싱해 note를 수집.
    - id는 ffill
    - 각 id에 대해 '비어있지 않은 첫 note'를 채택
    """
    if "id" not in df.columns or "metadata" not in df.columns:
        return {}

    tmp = df.copy()
    tmp["id"] = tmp["id"].ffill().astype(str)

    out: Dict[str, str] = {}
    for _, row in tmp.iterrows():
        _id = row["id"].strip()
        if not _id:
            continue
        meta_dict = _parse_metadata_cell(row["metadata"])
        note = str(meta_dict.get("note", "") or "").strip()
        if note and _id not in out:
            out[_id] = note
    return out

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")
LINK_BLUE = "0563C1"

# [타입] 문장 형태 파싱용 ([Type] 내용)
TYPE_BRACKET_RE = re.compile(r"^\s*\[(.+?)\]\s*(.*)$")


# =========================
# JSON -> Excel (정방향)
# =========================
def extract_sentences(doc: Dict[str, Any]) -> List[Tuple[str, str]]:
    """
    EX.exp_sentence 내부를 탐색해 ([Type] sentence) 또는 그냥 sentence를
    (type, sentence) 튜플 리스트로 반환
    """
    out: List[Tuple[str, str]] = []
    for ex in doc.get("EX", []):
        for item in ex.get("exp_sentence", []) or []:
            if not isinstance(item, dict):
                continue
            for _k, v in item.items():
                seq = v if isinstance(v, list) else [v]
                for s in seq:
                    if not s:
                        continue
                    text = str(s).strip()
                    m = TYPE_BRACKET_RE.match(text)
                    if m:
                        out.append((m.group(1).strip(), m.group(2).strip()))
                    else:
                        out.append(("", text))
    return out


def _clean_url(u: str) -> str:
    if not u:
        return ""
    u = str(u).strip()
    if (u.startswith('"') and u.endswith('"')) or (u.startswith("'") and u.endswith("'")):
        u = u[1:-1].strip()
    return u


def format_metadata_and_url(meta: Dict[str, Any]) -> Tuple[str, str]:
    """
    metadata를 멀티라인 문자열로 정리하고, url만 분리해서 반환
    """
    url_only = _clean_url(meta.get("url", ""))
    lines = ['metadata : {']
    for k in META_ORDER:
        v = meta.get(k, "")
        if k == "url":
            v = url_only or meta.get("url", "")
        lines.append(f'  "{k}": "{v}",' if k != META_ORDER[-1] else f'  "{k}": "{v}"')
    lines.append("}")
    return "\n".join(lines), url_only


def extract_mdfcn_memo(mdfcn_infos):
    """
    mdfcn_infos[*].mdfcn_memo 가 JSON 문자열이면 value만 추출해
    "작업 목록 1 : ..." 형태로 번호 매겨 반환.
    항목 간에 빈 줄 한 줄 추가.
    """
    if not mdfcn_infos:
        return ""
    out, idx = [], 1
    for info in mdfcn_infos:
        raw = info.get("mdfcn_memo", "")
        if not raw:
            continue
        try:
            arr = json.loads(raw)
            if isinstance(arr, list):
                for obj in arr:
                    val = str((obj or {}).get("value", "")).strip()
                    if val:
                        out.append(f"작업 목록 {idx} : {val}")
                        idx += 1
        except Exception:
            val = str(raw).strip()
            if val:
                out.append(f"작업 목록 {idx} : {val}")
                idx += 1
    return "\n\n".join(out)


def to_rows(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    JSON(dict) -> 행 리스트
    """
    rows: List[Dict[str, Any]] = []
    docs = data.get("document", [])
    if not isinstance(docs, list):
        return rows

    for doc in docs:
        img_id = str(doc.get("id", ""))
        meta = doc.get("metadata", {}) or {}
        medium_category = str(meta.get("Medium_category", "") or "")
        worker_id_cnst = str(doc.get("worker_id_cnst", "") or "")

        md_text, md_url = format_metadata_and_url(meta)
        memo_text = extract_mdfcn_memo(doc.get("mdfcn_infos", []) or [])

        pairs = extract_sentences(doc) or [("", "")]
        for typ, sent in pairs:
            rows.append({
                "id": img_id,
                "worker_id_cnst": worker_id_cnst,
                "Medium_category": medium_category,
                "유형": typ,
                "설명 문장": sent,
                "metadata": md_text,
                "meta_url": md_url,  # 하이퍼링크용(엑셀 열에는 포함 안 함)
                "mdfcn_memo(검수자 수정 이력)": memo_text,
            })
    return rows


def estimate_wrapped_lines(text: str, col_chars: int) -> int:
    if not text:
        return 1
    total = 0
    width = max(col_chars, 5)
    for para in str(text).split("\n"):
        total += max(1, math.ceil(len(para) / (width * 1.08)))
    return max(1, total)


def _write_excel_to_bytes(all_rows: List[Dict[str, Any]]) -> bytes:
    """
    행 리스트 -> Excel bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "result"

    headers = [
        "id", "worker_id_cnst", "Medium_category",
        "유형", "설명 문장", "metadata", "mdfcn_memo\n(검수자 수정 이력)"
    ]
    ws.append(headers)

    # 헤더 스타일
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        cell.fill = HEADER_FILL

    # 열 너비
    widths = {1: 12, 2: 16, 3: 14, 4: 16, 5: 80, 6: 60, 7: 50}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 그룹 시작/개수 추적
    start_row_by_group: Dict[Tuple[str], int] = {}
    count_by_group: Dict[Tuple[str], int] = {}

    current_row = 2
    for row in all_rows:
        ws.append([
            xls_safe(row.get("id", "")),
            xls_safe(row.get("worker_id_cnst", "")),
            xls_safe(row.get("Medium_category", "")),
            xls_safe(row.get("유형", "")),
            xls_safe(row.get("설명 문장", "")),
            xls_safe(row.get("metadata", "")),
            xls_safe(row.get("mdfcn_memo(검수자 수정 이력)", "")),
        ])
        for c in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=c).alignment = Alignment(
                vertical="top", wrap_text=(c in (5, 6, 7))
            )
            ws.cell(row=current_row, column=c).border = THIN_BORDER

        key = (row.get("id",""),)
        if key not in start_row_by_group:
            start_row_by_group[key] = current_row
            count_by_group[key] = 0
        count_by_group[key] += 1
        current_row += 1

    # 병합
    merge_cols = [1, 2, 3, 6, 7]
    for key, start in start_row_by_group.items():
        cnt = count_by_group[key]
        if cnt > 1:
            end = start + cnt - 1
            for col in merge_cols:
                ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)
                ws.cell(row=start, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    # metadata 하이퍼링크(같은 id 첫 행만)
    first_url_by_id: Dict[str, str] = {}
    for _r, row in enumerate(all_rows, start=2):
        rid = row.get("id", "")
        if rid and rid not in first_url_by_id:
            first_url_by_id[rid] = row.get("meta_url", "") or ""

    from openpyxl.cell.cell import MergedCell
    try:
        from openpyxl.worksheet.hyperlink import Hyperlink
    except Exception:
        Hyperlink = None

    for key, start in start_row_by_group.items():
        doc_id = key[0]
        url = str(first_url_by_id.get(doc_id, "") or "").strip()
        if not (url and url.startswith(("http://", "https://"))):
            continue

        c = ws.cell(row=start, column=6)
        if isinstance(c, MergedCell):
            if Hyperlink is not None:
                ws._hyperlinks.append(Hyperlink(ref=c.coordinate, target=url, display=url))
        else:
            try:
                c.hyperlink = url
            except AttributeError:
                if Hyperlink is not None:
                    ws._hyperlinks.append(Hyperlink(ref=c.coordinate, target=url, display=url))

        c.font = Font(color=LINK_BLUE, underline="none")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = THIN_BORDER

    # 행 높이 조정
    LINE_HEIGHT_PT = 18
    group_starts = set(start_row_by_group.values())
    for r in range(2, current_row):
        desc = ws.cell(row=r, column=5).value or ""
        desc_lines = estimate_wrapped_lines(desc, widths[5])
        if r in group_starts:
            meta_plain = ws.cell(row=r, column=6).value or ""
            memo_plain = ws.cell(row=r, column=7).value or ""
            need = max(
                desc_lines,
                estimate_wrapped_lines(meta_plain, widths[6]),
                estimate_wrapped_lines(memo_plain, widths[7]),
            )
        else:
            need = desc_lines
        ws.row_dimensions[r].height = max(1, need) * LINE_HEIGHT_PT

    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def photo_json_to_xlsx_bytes(data: Dict[str, Any]) -> bytes:
    rows = to_rows(data)
    if not rows:
        return _write_excel_to_bytes([])
    return _write_excel_to_bytes(rows)

def _read_excel_multi(ef, sheet_name: Optional[Iterable[str] or str] = None) -> pd.DataFrame:
    """
    Excel 파일에서 시트를 읽어 하나의 DataFrame으로 합친다.
    """
    need_cols = ["id", "설명 문장"]
    opt_cols  = ["유형", "Medium_category"]

    if sheet_name is None:
        sheets = pd.read_excel(ef, sheet_name=None)
        dfs = []
        for name, df in sheets.items():
            df = df.copy()
            for c in need_cols + opt_cols:
                if c not in df.columns:
                    df[c] = ""
            df["__sheet__"] = str(name)
            dfs.append(df)
        if not dfs:
            return pd.DataFrame(columns=need_cols + opt_cols)
        out = pd.concat(dfs, ignore_index=True)
        return out

    if isinstance(sheet_name, str):
        df = pd.read_excel(ef, sheet_name=sheet_name)
        df = df.copy()
        for c in need_cols + opt_cols:
            if c not in df.columns:
                df[c] = ""
        df["__sheet__"] = str(sheet_name)
        return df

    try:
        names = list(sheet_name)
    except TypeError:
        raise TypeError("sheet_name은 None, 문자열, 또는 문자열 리스트여야 합니다.")
    dfs = []
    all_sheets = pd.read_excel(ef, sheet_name=None)
    for nm in names:
        if nm not in all_sheets:
            continue
        df = all_sheets[nm].copy()
        for c in need_cols + opt_cols:
            if c not in df.columns:
                df[c] = ""
        df["__sheet__"] = str(nm)
        dfs.append(df)
    if not dfs:
        return pd.DataFrame(columns=need_cols + opt_cols)
    return pd.concat(dfs, ignore_index=True)

# ==========================================
# Excel('설명 문장') → JSON (역방향, ZIP 지원)
# ==========================================

def _cleanup_exp_sentences(doc: Dict[str, Any]) -> None:
    """
    빈 문자열/빈 리스트/빈 딕셔너리를 걷어내서 exp_sentence 구조를 정리.
    """
    ex_list = doc.get("EX", [])
    if not isinstance(ex_list, list):
        return

    for ex in ex_list:
        exp = ex.get("exp_sentence")
        if exp is None:
            continue

        if isinstance(exp, list):
            new_exp = []
            for item in exp:
                if isinstance(item, dict):
                    new_item = {}
                    for k, v in item.items():
                        if isinstance(v, list):
                            vv = [str(s).strip() for s in v if str(s or "").strip()]
                            if vv:
                                new_item[k] = vv
                        else:
                            s = str(v or "").strip()
                            if s:
                                new_item[k] = s
                    if new_item:
                        new_exp.append(new_item)
            ex["exp_sentence"] = new_exp
            if not new_exp:
                ex.pop("exp_sentence", None)

        elif isinstance(exp, dict):
            new_exp = {}
            for k, v in exp.items():
                if isinstance(v, list):
                    vv = [str(s).strip() for s in v if str(s or "").strip()]
                    if vv:
                        new_exp[k] = vv
                else:
                    s = str(v or "").strip()
                    if s:
                        new_exp[k] = s
            if new_exp:
                ex["exp_sentence"] = new_exp
            else:
                ex.pop("exp_sentence", None)


def _compose_text_with_type(old_text: str, new_sentence: str, excel_type: str) -> str:
    """
    엑셀 '유형' + '설명 문장'을 하나의 문자열로 합치기.
    이 버전에서는 old_text는 무시하고, 항상 엑셀 기준으로 만든다고 보면 된다.
    """
    s_new = "" if new_sentence is None else str(new_sentence).strip()
    t_new = "" if excel_type is None else str(excel_type).strip()

    # 엑셀 유형에 [ ]가 이미 붙어 있으면 제거
    if t_new.startswith("[") and t_new.endswith("]"):
        t_new = t_new[1:-1].strip()

    body = s_new

    if t_new:
        return f"[{t_new}] {body}".strip() if body else f"[{t_new}]"
    else:
        return body


def _collect_excel_pairs_by_id(df: pd.DataFrame, skip_blank: bool = True) -> Dict[str, List[Tuple[str, str]]]:
    """
    엑셀에서 id별 (유형, 설명 문장) 시퀀스를 원본 행 순서대로 수집.
    - 병합 셀로 인해 비는 id/유형은 ffill로 채움
    - skip_blank=True면 빈 '설명 문장'은 건너뜀
    반환: { id: [(type, sentence), ...], ... }
    """
    required = {"id", "설명 문장"}
    if not required.issubset(set(df.columns)):
        raise ValueError("엑셀에 'id', '설명 문장' 컬럼이 필요합니다.")

    tmp = df.copy()
    if "id" in tmp.columns:
        tmp["id"] = tmp["id"].ffill()
    if "유형" in tmp.columns:
        tmp["유형"] = tmp["유형"].ffill()

    tmp["id"] = tmp["id"].astype(str)
    if "유형" not in tmp.columns:
        tmp["유형"] = ""
    tmp["유형"] = tmp["유형"].fillna("").astype(str)
    tmp["설명 문장"] = tmp["설명 문장"].fillna("").astype(str)

    bucket: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
    for _, row in tmp.iterrows():
        _id = row["id"].strip()
        typ = row["유형"].strip()
        sent = row["설명 문장"].strip()
        if skip_blank and not sent:
            continue
        bucket[_id].append((typ, sent))
    return bucket


def _collect_medium_by_id(df: pd.DataFrame) -> Dict[str, str]:
    """
    엑셀에서 id별 Medium_category 값을 수집.
    - 병합 셀 보정을 위해 id, Medium_category ffill
    - 각 id에 대해 '비어있지 않은 첫 값'을 채택
    """
    if "id" not in df.columns:
        return {}

    tmp = df.copy()
    tmp["id"] = tmp["id"].ffill().astype(str)

    if "Medium_category" not in tmp.columns:
        return {}

    tmp["Medium_category"] = tmp["Medium_category"].ffill().fillna("").astype(str)

    out: Dict[str, str] = {}
    for _, row in tmp.iterrows():
        _id = row["id"].strip()
        mc = row["Medium_category"].strip()
        if _id and mc and _id not in out:
            out[_id] = mc
    return out


def apply_excel_desc_to_photo_json(
    json_obj: Dict[str, Any],
    excel_df: pd.DataFrame,
    skip_blank: bool = False
) -> Dict[str, Any]:
    """
    사진 JSON에 엑셀의 '설명 문장'과 'Medium_category'를 반영.

    ★ 규칙 (요청사항에 맞게 변경) ★
    1) 엑셀에 없는 id → JSON 설명문장은 그대로 둔다.
    2) 엑셀에 해당 id가 있고, 유효한 '설명 문장'이 하나라도 있으면:
       - 기존 JSON의 설명문장 슬롯을 전부 무시하고,
       - 엑셀의 (유형, 설명 문장) 시퀀스로 '완전히 새로' exp_sentence를 구성해 덮어쓴다.
       - JSON에 원래 설명문장이 없던 경우도 동일하게 새로 만든다.
    """

    mapping = _collect_excel_pairs_by_id(excel_df, skip_blank=skip_blank)
    metadata_map = _collect_metadata_by_id(excel_df)
    medium_map = _collect_medium_by_id(excel_df)
    note_map = _collect_note_by_id(excel_df)

    docs = json_obj.get("document", [])
    if not isinstance(docs, list):
        return json_obj

    for doc in docs:
        doc_id = str(doc.get("id", ""))

        # metadata / Medium_category / note는 id만 맞으면 그대로 반영
        if doc_id and (doc_id in metadata_map or doc_id in medium_map or doc_id in note_map or not skip_blank):
            if not isinstance(doc.get("metadata"), dict):
                doc["metadata"] = {}
            meta_obj = doc["metadata"]
        else:
            meta_obj = None

        if meta_obj is not None:
            meta_from_excel = metadata_map.get(doc_id)
            if meta_from_excel:
                meta_obj.update(meta_from_excel)

            mc_val = medium_map.get(doc_id, "")
            if mc_val or not skip_blank:
                meta_obj["Medium_category"] = mc_val

            note_val = note_map.get(doc_id, "")
            if note_val or not skip_blank:
                meta_obj["note"] = note_val

        # --- 여기서부터 '설명 문장' 처리 ---

        # 엑셀에 이 id가 아예 없으면: 설명문장은 손대지 않고 그대로 둔다.
        if doc_id not in mapping:
            continue

        seq = mapping.get(doc_id, [])
        # 엑셀에 id는 있지만, 유효한 설명 문장이 하나도 없으면 역시 그대로 둔다.
        if not seq:
            continue

        # EX 리스트 확보 (없으면 새로 생성)
        ex_list = doc.get("EX")
        if not isinstance(ex_list, list) or not ex_list:
            doc["EX"] = [{}]
            ex_list = doc["EX"]

        # 기존 모든 EX의 exp_sentence는 제거 (완전 덮어쓰기)
        for ex in ex_list:
            if isinstance(ex, dict) and "exp_sentence" in ex:
                ex.pop("exp_sentence", None)

        # 첫 번째 EX에 새 exp_sentence 구성
        first_ex = ex_list[0]
        new_exp_list: List[Dict[str, Any]] = []

        for idx, (typ, sent) in enumerate(seq, 1):
            typ_clean = (typ or "").strip()
            sent_clean = (sent or "").strip()
            if not (typ_clean or sent_clean):
                # 둘 다 공백이면 무시
                continue

            composed = _compose_text_with_type("", sent_clean, typ_clean)  # → "[유형] 문장" 또는 "문장"
            key_name = f"설명문장{idx}"
            new_exp_list.append({key_name: [composed]})

        if new_exp_list:
            first_ex["exp_sentence"] = new_exp_list
        else:
            # seq가 있었지만 다 공백이었을 경우: 아무 것도 안 남기고 넘어감
            # (기존 exp_sentence는 이미 지운 상태이므로 '빈 상태'가 됨)
            pass

        _cleanup_exp_sentences(doc)

    return json_obj


def apply_excel_desc_to_json_from_zip(
    zip_bytes: bytes,
    sheet_name: Optional[str] = None,
    skip_blank: bool = True,
) -> Tuple[bytes, str]:
    """
    (사진 전용) ZIP(엑셀+단일 JSON)을 받아 엑셀의 '설명 문장'을 JSON에 반영.
    반환: (updated_json_bytes, suggested_filename)
    """
    if not isinstance(zip_bytes, (bytes, bytearray)):
        raise TypeError("zip_bytes는 bytes/bytearray여야 합니다.")

    with zipfile.ZipFile(BytesIO(zip_bytes), "r") as zf:
        json_members = [m for m in zf.namelist() if m.lower().endswith(".json")]
        xlsx_members = [m for m in zf.namelist() if m.lower().endswith(".xlsx")]
        xls_members  = [m for m in zf.namelist() if m.lower().endswith(".xls")]

        json_member = None
        for m in json_members:
            if Path(m).name.startswith("project_"):
                json_member = m
                break
        if json_member is None and json_members:
            json_member = json_members[0]

        excel_member = xlsx_members[0] if xlsx_members else (xls_members[0] if xls_members else None)

        if not json_member:
            raise FileNotFoundError("ZIP 안에 JSON 파일이 없습니다.")
        if not excel_member:
            raise FileNotFoundError("ZIP 안에 Excel 파일(.xlsx/.xls)이 없습니다.")

        with zf.open(json_member) as jf:
            json_obj = json.loads(jf.read().decode("utf-8"))

        with zf.open(excel_member) as ef:
            df = _read_excel_multi(ef, sheet_name=sheet_name)

        updated = apply_excel_desc_to_photo_json(json_obj, df, skip_blank=skip_blank)

        base = Path(json_member).name
        out_name = (base[:-5] if base.lower().endswith(".json") else base) + "_updated.json"

        text = json.dumps(updated, ensure_ascii=False, indent=2)
        return text.encode("utf-8"), out_name
