#dataly_tools/table_to_excel.py

# -*- coding: utf-8 -*-
"""
표(JSON) -> Excel 변환기 (bytes 반환)
- 입력: dict(JSON 파싱 결과)
- 출력: bytes(XLSX)
- 같은 id 블록 기준으로 [A:id, D:metadata] 병합
- metadata 첫 행에만 URL 하이퍼링크
"""
import json
from collections import defaultdict
from io import BytesIO
from typing import Any, Dict, Iterable, List, Tuple, Optional

import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

import zipfile
from pathlib import Path

import unicodedata as ud

def _norm_colname(s: str) -> str:
    if s is None:
        return ""
    s = ud.normalize("NFC", str(s))
    s = s.replace("\u200b", "").replace("\ufeff", "")  # 제로폭, BOM
    s = s.replace("\xa0", " ")  # NBSP → 스페이스
    base = s.strip()
    tight = base.replace(" ", "").lower()

    # 표준화 매핑
    if tight in ("id", "아이디"):
        return "id"
    if tight in ("유형", "type", "referencetype", "reference", "참조유형"):
        return "유형"
    if tight in ("설명문장", "설명문", "설명", "expsentence", "exp_sentence"):
        return "설명 문장"
    # 그 외는 원래(정리된) 이름
    return base

def _normalize_excel_columns(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns={c: _norm_colname(c) for c in df.columns})

def _norm_key(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = ud.normalize("NFC", s)
    s = s.replace("\u200b", "").replace("\ufeff", "")  # 제로폭/ BOM 제거
    return s.strip()

def _norm_ref_type(s: str) -> str:
    if not s:
        return ""
    s = ud.normalize("NFC", str(s)).strip().lower()
    # 다양한 표기 → 밑줄 표기 통일
    s = s.replace("-", "_").replace(" ", "_")
    # tableRef, tableReference 등도 커버
    s = s.replace("tableref", "table_ref").replace("rowref", "row_ref").replace("colref", "col_ref").replace("cellref", "cell_ref")
    return s

# ===== 유형 매핑 =====
REF_MAP = {
    "table_ref": "표 설명 문장",
    "row_ref": "행 설명 문장",
    "col_ref": "열 설명 문장",
    "cell_ref": "불연속 영역 설명 문장",
}
TYPE_TAGS = {"table_ref", "row_ref", "col_ref", "cell_ref"}

# 역매핑: 엑셀 '유형' → JSON reference_type
REF_MAP_INV = {v: k for k, v in REF_MAP.items()}


def _set_exp_sentence_on_dict(d: Dict[str, Any], new_sentence: str, prefer_existing: bool = True) -> None:
    """
    d(dict) 내부의 '설명문장' 계열 키(공백/변형 포함)를 모두 정리하고 하나의 키로만 기록.
    - prefer_existing=True: 기존에 쓰던 키명이 있으면 그 키를 유지하여 overwrite
    - 기존 키가 없다면 기본 키 '설명문장'으로 기록
    - 기록 형태는 호환성을 위해 list[str] 유지
    """
    if not isinstance(d, dict):
        return

    # 후보 키 수집(공백 제거 후 비교)
    candidates = []
    for k in list(d.keys()):
        kn = str(k).replace(" ", "")
        if kn in ("설명문장", "설명문장들", "설명문", "설명"):
            candidates.append(k)

    # 사용할 타깃 키 결정
    target_key = candidates[0] if (prefer_existing and candidates) else "설명문장"

    # 중복 방지: 기존 후보 키 제거
    for k in candidates:
        try:
            del d[k]
        except Exception:
            pass

    d[target_key] = ["" if new_sentence is None else str(new_sentence)]

def _iter_exp_items(ex_obj):
    """ex_obj.get('exp_sentence')가 list/dict/str 어떤 형태든 리스트로 정규화"""
    raw = ex_obj.get("exp_sentence", [])
    if isinstance(raw, list):
        return raw
    if isinstance(raw, dict):
        return [raw]
    if isinstance(raw, str):
        return [{"설명문장": [raw]}]
    return []

def _pick_sentence(exp_item) -> str:
    """
    exp_item에서 설명문장 1개를 뽑아 반환.
    - '설명문장' / '설명 문장' / 비슷한 변형(공백 제거 후 비교) 우선
    - 그래도 없으면 dict 값들 중 list[str] 또는 str을 첫 번째로 사용
    """
    if isinstance(exp_item, str):
        return exp_item.strip()

    if isinstance(exp_item, dict):
        # 1) 우선 '설명문장' / '설명 문장' 같이 보이는 키를 탐색(공백 제거 후 비교)
        for k, v in exp_item.items():
            kn = str(k).replace(" ", "")
            if kn in ("설명문장", "설명문장들", "설명"):
                if isinstance(v, list) and v:
                    return str(v[0]).strip()
                return str(v).strip() if v is not None else ""

        # 2) fallback: 값들 중 list[str] 또는 str을 사용
        for v in exp_item.values():
            if isinstance(v, list) and v and isinstance(v[0], str):
                return v[0].strip()
            if isinstance(v, str):
                return v.strip()

    return ""  # 못 찾으면 빈 문자열


def extract_mdfcn_values(obj, sep: str = "\n") -> str:
    """mdfcn_infos에서 value만 추출(중복 제거, 순서 유지) 후 sep로 결합"""
    values: List[str] = []

    def _dedup_keep_order(items: List[str]) -> List[str]:
        seen, out = set(), []
        for it in items:
            if it and it not in seen:
                seen.add(it)
                out.append(it)
        return out

    def _walk(x):
        if x is None:
            return
        if isinstance(x, str):
            s = x.strip()
            if not s:
                return
            if s[:1] in ("[", "{"):
                try:
                    _walk(json.loads(s))
                    return
                except Exception:
                    if s not in TYPE_TAGS:
                        values.append(s)
                    return
            if s not in TYPE_TAGS:
                values.append(s)
            return
        if isinstance(x, dict):
            v = x.get("value")
            if isinstance(v, str):
                v = v.strip()
                if v:
                    values.append(v)
            mm = x.get("mdfcn_memo")
            if isinstance(mm, str):
                mm_s = mm.strip()
                if mm_s:
                    try:
                        _walk(json.loads(mm_s))
                    except Exception:
                        pass
            for k, sub in x.items():
                if k in ("value", "mdfcn_memo"):
                    continue
                if isinstance(sub, (list, dict)):
                    _walk(sub)
            return
        if isinstance(x, (list, tuple)):
            for it in x:
                _walk(it)
            return

    _walk(obj)
    return sep.join(_dedup_keep_order([v for v in values if isinstance(v, str)]))


def extract_url(meta: Any) -> str:
    """metadata.url을 안전하게 추출"""
    if isinstance(meta, dict):
        u = meta.get("url", "")
        if isinstance(u, list):
            return u[0] if u else ""
        return str(u or "")
    return ""


def table_json_to_xlsx_bytes(data: Dict[str, Any]) -> bytes:
    """datalyManager에서 호출하는 공개 API
    - worker_id_cnst, mdfcn_infos 컬럼 제거 버전
    """
    rows: List[Dict[str, Any]] = []
    group_counts = defaultdict(int)

    for doc in data.get("document", []) or []:
        doc_id = doc.get("id", "")
        metadata = doc.get("metadata", {}) or {}

        # URL 추출
        url = extract_url(metadata)

        for ex in doc.get("EX", []) or []:
            ref_type = ex.get("reference", {}).get("reference_type", "")
            for exp_item in _iter_exp_items(ex):
                sentence = _pick_sentence(exp_item)  # ← 키 변형 안전 처리
                rows.append({
                    "id": doc_id,
                    "유형": REF_MAP.get(ref_type, ref_type),
                    "설명 문장": sentence,
                    "metadata": json.dumps(metadata, ensure_ascii=False, indent=2),
                    "url": url,
                })
                group_counts[doc_id] += 1

    # 빈 데이터여도 헤더만 있는 파일 생성
    df = pd.DataFrame(
        rows,
        columns=["id", "유형", "설명 문장", "metadata"]
    ).fillna("")

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="sheet1")
        ws = writer.sheets["sheet1"]

        # 열 너비 (A: id, B: 유형, C: 설명 문장, D: metadata)
        widths = {"A": 18, "B": 16, "C": 80, "D": 50}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # 머리글 스타일
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # metadata 헤더 강조
        ws["D1"].fill = PatternFill("solid", fgColor="BDD7EE")

        # 데이터 영역: 줄바꿈 + 상단 정렬 + 테두리
        thin = Side(style="thin", color="999999")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        max_row = ws.max_row
        max_col = ws.max_column

        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if c >= 3:  # 설명 문장, metadata
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(vertical="top")
                cell.border = border

        # id 블록 병합: A(id), D(metadata)
        cur_row = 2
        for doc_id, count in group_counts.items():
            if count > 1:
                for col in (1, 4):  # A, D
                    ws.merge_cells(
                        start_row=cur_row, start_column=col,
                        end_row=cur_row + count - 1, end_column=col
                    )
                    top_cell = ws.cell(row=cur_row, column=col)
                    top_cell.alignment = Alignment(vertical="top", wrap_text=True)
                    for rr in range(cur_row, cur_row + count):
                        ws.cell(row=rr, column=col).border = border
            cur_row += count

        # 하이퍼링크: 같은 id의 첫 행만 D열(metadata)에 설정
        first_row_for_id: Dict[str, int] = {}
        for idx, row in enumerate(rows, start=2):
            first_row_for_id.setdefault(row["id"], idx)

        for idx, row in enumerate(rows, start=2):
            url = row.get("url", "")
            if not url:
                continue
            if idx != first_row_for_id[row["id"]]:
                continue
            ws.cell(row=idx, column=4).hyperlink = url  # D열

        # 행 높이: C/D의 개행 수 기준 근사 조절
        for r in range(2, max_row + 1):
            max_lines = 1
            for c in (3, 4):  # C, D
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and "\n" in val:
                    lines = val.count("\n") + 1
                    if lines > max_lines:
                        max_lines = lines
            ws.row_dimensions[r].height = min(15 + (max_lines - 1) * 12, 200)

        # 틀 고정
        ws.freeze_panes = "A2"

    output.seek(0)
    return output.getvalue()



def _collect_excel_sentences_by_id(df: pd.DataFrame) -> Dict[str, List[str]]:
    if "id" not in df.columns or "설명 문장" not in df.columns:
        raise ValueError("엑셀에 'id'와 '설명 문장' 컬럼이 필요합니다.")
    tmp = df.copy()
    tmp["id"] = tmp["id"].ffill().astype(str).map(_norm_key)
    tmp["설명 문장"] = tmp["설명 문장"].fillna("").astype(str)

    bucket: Dict[str, List[str]] = defaultdict(list)
    for _, row in tmp.iterrows():
        _id = row["id"]
        sent = row["설명 문장"].strip()
        bucket[_id].append(sent)
    return bucket



def _iter_exp_slots(ex_obj):
    """
    ex_obj 내부의 exp_sentence '슬롯'을 순서대로 반환.
    각 슬롯은 ('list'|'dict'|'str', container, index_or_key) 형태로 기술.
    - list: (mode='list', list_obj, idx)
    - dict: (mode='dict', dict_obj, None) → dict['설명문장']에 기록
    - str : (mode='str',  ex_obj(부모), 'exp_sentence') → ex_obj['exp_sentence'] 문자열에 기록
    기타 타입/누락 시, 기본 슬롯을 생성해서 반환.
    """
    if not isinstance(ex_obj, dict):
        return  # 방어적 처리

    if "exp_sentence" not in ex_obj or ex_obj["exp_sentence"] is None:
        # 기본 한 슬롯을 만들어 준다(문자열)
        ex_obj["exp_sentence"] = ""
        yield ("str", ex_obj, "exp_sentence")
        return

    raw = ex_obj["exp_sentence"]

    if isinstance(raw, list):
        # 각 원소가 dict/str 등 무엇이든 슬롯으로 간주
        for i in range(len(raw)):
            yield ("list", raw, i)
        return

    if isinstance(raw, dict):
        # dict 한 덩어리를 하나의 슬롯으로 간주: dict['설명문장']에 기록
        yield ("dict", raw, None)
        return

    if isinstance(raw, str):
        # 문자열 하나면 그 자체가 슬롯
        yield ("str", ex_obj, "exp_sentence")
        return

    # 알 수 없는 타입이면 문자열 슬롯으로 강제
    ex_obj["exp_sentence"] = ""
    yield ("str", ex_obj, "exp_sentence")


def _assign_sentence_to_slot(slot, new_sentence: str):
    """
    슬롯 정의에 따라 new_sentence를 해당 위치에 기록.
    dict/list-dict 슬롯은 _set_exp_sentence_on_dict()로 정규화하여
    '설명문장' 계열 키가 둘 이상 생기지 않도록 하나의 키로만 overwrite.
    """
    mode, container, pos = slot
    s = "" if new_sentence is None else str(new_sentence)

    if mode == "list":
        item = container[pos]
        if isinstance(item, dict):
            # 기존/변형 키 정리 후 하나의 키로 overwrite
            _set_exp_sentence_on_dict(item, s, prefer_existing=True)
        else:
            container[pos] = s
        return

    if mode == "dict":
        # 기존/변형 키 정리 후 하나의 키로 overwrite
        _set_exp_sentence_on_dict(container, s, prefer_existing=True)
        return

    if mode == "str":
        container[pos] = s
        return


def apply_excel_desc_to_json(json_obj: Dict[str, Any], excel_df: pd.DataFrame) -> Dict[str, Any]:
    """
    엑셀의 '설명 문장' 값을 JSON의 exp_sentence에 반영하여 JSON 객체를 반환.
    매핑 규칙:
      - 같은 id 블록 내에서 행 순서대로 EX의 exp_sentence '슬롯'에 순차 매핑
      - 슬롯 수 > 엑셀 문장 수 → 남는 슬롯은 원본 유지
      - 슬롯 수 < 엑셀 문장 수 → 초과 문장은 무시
    """
    mapping = _collect_excel_sentences_by_id(excel_df)

    docs = json_obj.get("document", [])
    if not isinstance(docs, list):
        return json_obj  # 형식 방어

    for doc in docs:
        doc_id = _norm_key(doc.get("id", ""))
        seq = mapping.get(doc_id, [])
        if not seq:
            continue

        used = 0
        ex_list = doc.get("EX", [])
        if not isinstance(ex_list, list):
            continue

        for ex in ex_list:
            for slot in _iter_exp_slots(ex):
                if used >= len(seq):
                    break
                _assign_sentence_to_slot(slot, seq[used])
                used += 1

            if used >= len(seq):
                break

    return json_obj

def _read_excel_multi(ef, sheet_name: Optional[Iterable[str] or str] = None) -> pd.DataFrame:
    """
    Excel 파일에서 시트를 읽어 하나의 DataFrame으로 합침.
    - 모든 시트/지정 시트 지원
    - 컬럼명을 id / 유형 / 설명 문장으로 정규화
    - 누락 컬럼은 빈 컬럼으로 보정
    """
    need_cols = ["id", "유형", "설명 문장"]  # 유형은 없어도 동작하지만, 여기선 기본 세트로 맞춤

    def _prep(df: pd.DataFrame, name: str) -> pd.DataFrame:
        df = _normalize_excel_columns(df.copy())
        for c in need_cols:
            if c not in df.columns:
                df[c] = ""
        df["__sheet__"] = str(name)
        return df

    if sheet_name is None:
        sheets = pd.read_excel(ef, sheet_name=None)
        dfs = []
        for name, df in sheets.items():
            dfs.append(_prep(df, name))
        if not dfs:
            return pd.DataFrame(columns=need_cols + ["__sheet__"])
        return pd.concat(dfs, ignore_index=True)

    if isinstance(sheet_name, str):
        df = pd.read_excel(ef, sheet_name=sheet_name)
        return _prep(df, sheet_name)

    # 시트명 리스트
    names = list(sheet_name)
    all_sheets = pd.read_excel(ef, sheet_name=None)
    dfs = []
    for nm in names:
        if nm in all_sheets:
            dfs.append(_prep(all_sheets[nm], nm))
    if not dfs:
        return pd.DataFrame(columns=need_cols + ["__sheet__"])
    return pd.concat(dfs, ignore_index=True)

def _pick_zip_members(zf: zipfile.ZipFile):
    """
    ZIP에서 JSON 1개, XLSX 1개를 추출 대상으로 선택.
    - JSON은 project_*.json 우선, 그 외 첫 번째 .json
    - Excel은 .xlsx 우선(.xls는 의존성에 따라 미지원일 수 있음)
    """
    json_members = [m for m in zf.namelist() if m.lower().endswith(".json")]
    xlsx_members = [m for m in zf.namelist() if m.lower().endswith(".xlsx")]
    xls_members  = [m for m in zf.namelist() if m.lower().endswith(".xls")]

    # JSON 우선순위: project_*
    json_member = None
    for m in json_members:
        name = Path(m).name
        if name.startswith("project_"):
            json_member = m
            break
    if json_member is None and json_members:
        json_member = json_members[0]

    # Excel 우선순위: .xlsx → (가능하면 .xls 대체)
    excel_member = xlsx_members[0] if xlsx_members else (xls_members[0] if xls_members else None)

    return json_member, excel_member

def _collect_excel_sentences_by_id_type(df: pd.DataFrame, skip_blank: bool = False) -> Dict[str, Dict[str, List[str]]]:
    required = {"id", "유형", "설명 문장"}
    if not required.issubset(set(df.columns)):
        raise ValueError("엑셀에 'id', '유형', '설명 문장' 컬럼이 모두 필요합니다.")

    tmp = df.copy()
    tmp["id"] = tmp["id"].ffill().astype(str).map(_norm_key)
    tmp["유형"] = tmp["유형"].ffill().astype(str)
    tmp["설명 문장"] = tmp["설명 문장"].fillna("").astype(str)

    bucket: Dict[str, Dict[str, List[str]]] = defaultdict(lambda: defaultdict(list))
    for _, row in tmp.iterrows():
        _id = row["id"]
        label = row["유형"].strip()
        sent = row["설명 문장"].strip()
        if skip_blank and not sent:
            continue
        ref_type = _label_to_ref_type(label)
        bucket[_id][ref_type].append(sent)
    return bucket


def _label_to_ref_type(label: Any) -> str:
    """
    엑셀 '유형' 라벨을 JSON reference_type 표준 값(table_ref/row_ref/col_ref/cell_ref)으로 정규화.
    - 기존 역매핑(REF_MAP_INV) 먼저 적용
    - 공백/밑줄 제거 후 커스텀 라벨 매핑 추가(대상 식별 문장/형태/색채/구성 요소/(비)역사)
    - 불연속* 시작어는 cell_ref 처리
    - 그 외에는 원문 라벨 반환(매칭 실패 시 상위 로직에서 폴백)
    """
    s = "" if label is None else str(label).strip()
    # 1차: 정확 매칭(기존 역매핑)
    if s in REF_MAP_INV:
        return REF_MAP_INV[s]

    # 2차: 공백/밑줄 제거 후 느슨 매칭
    s2 = s.replace(" ", "").replace("_", "")

    # 기존 규칙 + 커스텀 라벨 추가
    norm = {
        # 기존 약칭들
        "표설명문장": "table_ref", "표설명": "table_ref", "표": "table_ref",
        "행설명문장": "row_ref",   "행설명": "row_ref",   "행": "row_ref",
        "열설명문장": "col_ref",   "열설명": "col_ref",   "열": "col_ref",
        "불연속영역설명문장": "cell_ref", "불연속영역설명": "cell_ref",
        "불연속영역": "cell_ref", "불연속": "cell_ref",

        # ▼ 커스텀 라벨 매핑(필요에 맞게 조정 가능)
        "대상식별문장": "table_ref",
        "형태": "row_ref",
        "색채": "col_ref",
        "구성요소": "cell_ref",
        "(비)역사": "row_ref",
    }
    if s2 in norm:
        return norm[s2]

    # 3차: 시작어로 판별(예: '불연속영역 설명...'과 같은 변형)
    if s2.startswith("불연속"):
        return "cell_ref"

    return s  # 마지막 수단: 원문 라벨 그대로(매칭 실패 시 상위에서 폴백)


def apply_excel_desc_to_json(json_obj: Dict[str, Any], excel_df: pd.DataFrame, skip_blank: bool = True) -> Dict[str, Any]:
    """
    엑셀의 '설명 문장'을 JSON의 exp_sentence에 반영.
    - '유형' 컬럼이 있으면 id+유형별 정밀 매핑
      · reference_type 비었거나 매칭 실패 → 해당 id의 모든 유형 문장을 폴백으로 순서 배정
    - '유형' 컬럼이 없으면 id 순서대로 일괄 배분
    """
    docs = json_obj.get("document", [])
    if not isinstance(docs, list):
        return json_obj

    has_type_col = "유형" in excel_df.columns

    if has_type_col:
        # id는 정규화된 키로 생성됨
        mapping_by_type = _collect_excel_sentences_by_id_type(excel_df, skip_blank=skip_blank)

        for doc in docs:
            doc_id = _norm_key(doc.get("id", ""))   # 정규화
            type_map = mapping_by_type.get(doc_id, {})
            if not type_map:
                continue

            ex_list = doc.get("EX", [])
            if not isinstance(ex_list, list):
                continue

            # 유형별 소비 인덱스 + 폴백 시퀀스(flatten)
            from collections import defaultdict
            used_by_type: Dict[str, int] = defaultdict(int)
            fallback_key = "__fallback__"
            fallback_seq: List[str] = []
            for _k, _v in type_map.items():
                fallback_seq.extend(_v)

            for ex in ex_list:
                ref = ex.get("reference", {}) or {}
                ref_type = _norm_ref_type(ref.get("reference_type", ""))  # "" 가능
                seq = type_map.get(ref_type, [])

                # 타입 매칭 실패 또는 공란 → 폴백
                if not seq:
                    seq = fallback_seq
                    key_for_used = fallback_key
                else:
                    key_for_used = ref_type

                if not seq:
                    continue

                used = used_by_type.get(key_for_used, 0)
                for slot in _iter_exp_slots(ex):
                    if used >= len(seq):
                        break
                    _assign_sentence_to_slot(slot, seq[used])
                    used += 1
                used_by_type[key_for_used] = used

        return json_obj

    # === '유형' 컬럼이 없는 경우(구버전) ===
    mapping = _collect_excel_sentences_by_id(excel_df)  # 이쪽도 _norm_key 사용함

    for doc in docs:
        doc_id = _norm_key(doc.get("id", ""))
        seq = mapping.get(doc_id, [])
        if not seq:
            continue

        used = 0
        ex_list = doc.get("EX", [])
        if not isinstance(ex_list, list):
            continue

        for ex in ex_list:
            for slot in _iter_exp_slots(ex):
                if used >= len(seq):
                    break
                s = seq[used]
                if skip_blank and (s is None or str(s).strip() == ""):
                    used += 1
                    continue
                _assign_sentence_to_slot(slot, s)
                used += 1
            if used >= len(seq):
                break

    return json_obj




def apply_excel_desc_to_json_from_zip(
    zip_bytes: bytes,
    sheet_name: Optional[str] = None,
    skip_blank: bool = True,
) -> Tuple[bytes, str]:
    """
    ZIP 바이트(엑셀 + 단일 JSON)를 받아,
    엑셀의 '설명 문장'을 JSON에 반영한 뒤 (updated_json_bytes, suggested_filename)을 반환.
    - 엑셀에 '유형' 컬럼이 있으면 id+유형 정밀 매핑 사용
    - skip_blank=True면 엑셀의 빈 문자열은 원본을 덮어쓰지 않음
    """
    if not isinstance(zip_bytes, (bytes, bytearray)):
        raise TypeError("zip_bytes는 bytes이어야 합니다.")

    with zipfile.ZipFile(BytesIO(zip_bytes), "r") as zf:
        json_member, excel_member = _pick_zip_members(zf)
        if not json_member:
            raise FileNotFoundError("ZIP 안에 JSON 파일이 없습니다.")
        if not excel_member:
            raise FileNotFoundError("ZIP 안에 Excel(.xlsx) 파일이 없습니다.")

        # JSON 로드
        with zf.open(json_member) as jf:
            json_obj = json.loads(jf.read().decode("utf-8"))

        # Excel 로드
        with zf.open(excel_member) as ef:
            df = _read_excel_multi(ef, sheet_name=sheet_name)

        # 반영 (유형 컬럼이 있으면 정밀 매핑 분기 사용)
        updated = apply_excel_desc_to_json(json_obj, df, skip_blank=skip_blank)

        # 출력 파일명 제안
        base = Path(json_member).name
        out_name = (base[:-5] if base.lower().endswith(".json") else base) + "_updated.json"

        return json.dumps(updated, ensure_ascii=False, indent=2).encode("utf-8"), out_name
