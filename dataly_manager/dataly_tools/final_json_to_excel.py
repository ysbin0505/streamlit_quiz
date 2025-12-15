# -*- coding: utf-8 -*-
"""
사진(이미지) JSON -> Excel 변환기 (+ 엑셀 수정본을 JSON에 역반영)
- 입력: dict(JSON 파싱 결과)
- 출력: bytes(XLSX), datalyManager에서 st.download_button으로 바로 다운로드
- 같은 document.id 묶음에서 [id, worker_id_cnst, Medium_category, metadata, mdfcn_memo] 세로 병합
- metadata: 멀티라인 텍스트 + 같은 id 첫 행에만 URL 하이퍼링크(파란색, 밑줄 없음)

추가:
- apply_excel_desc_to_json_from_zip(zip_bytes, sheet_name=None, skip_blank=True)
  : ZIP(엑셀+단일 JSON)을 받아 엑셀의 '설명 문장'을 JSON에 반영해 반환

✅ 2025-12: exp_sentence 신형 구조 지원
- 신형:
  "exp_sentence": {
    "설명 문장1": {"feature": "[대상 식별 문장]", "sent": "...."},
    "설명 문장2": {"feature": "...", "sent": "..."}
  }
- 구형(기존):
  EX[].exp_sentence 내부에 "[Type] sentence" 혹은 "sentence" 문자열들
"""

import json
import math
import re
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple, Optional, Union
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 표시 순서(메타 키)
META_ORDER = [
    "note", "image", "copyright", "term_id", "Major_category",
    "title", "url", "Medium_category", "domain", "media_id",
    "publisher", "term", "source_id",
]

_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")
LINK_BLUE = "0563C1"

# [타입] 문장 형태 파싱용 ([Type] 내용)
TYPE_BRACKET_RE = re.compile(r"^\s*\[(.+?)\]\s*(.*)$")

META_NOTE_RE = re.compile(r'"note"\s*:\s*"(?P<note>.*?)"', re.DOTALL)


def xls_safe(val) -> str:
    """
    openpyxl이 허용하지 않는 XML 제어문자를 제거.
    숫자/None도 문자열로 안전 변환.
    """
    if val is None:
        return ""
    s = str(val)
    s = s.replace("\x00", "")
    s = _ILLEGAL_XML_RE.sub("", s)
    return s


def _parse_metadata_cell(cell_text: Any) -> Dict[str, Any]:
    """
    'metadata : { ... }' 형태의 멀티라인 문자열에서 { ... } 만 추출하여 json.loads 시도.
    엑셀에서 따옴표가 이중("...")으로 들어간 경우도 복원.
    실패 시 최소한 "note"만 정규식으로 추출.
    """
    if cell_text is None:
        return {}
    s = str(cell_text).strip()
    if not s:
        return {}

    i, j = s.find("{"), s.rfind("}")
    if i == -1 or j == -1 or i >= j:
        s_fix = s.replace('""', '"')
        m = META_NOTE_RE.search(s_fix)
        return {"note": m.group("note")} if m else {}

    blob = s[i:j + 1].strip()
    for candidate in (blob, blob.replace('""', '"')):
        try:
            return json.loads(candidate)
        except Exception:
            pass

    s_fix = blob.replace('""', '"')
    m = META_NOTE_RE.search(s_fix)
    return {"note": m.group("note")} if m else {}


def _collect_metadata_by_id(df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """
    엑셀 DF에서 id별로 metadata 셀을 파싱해 전체 metadata dict를 수집.
    - id는 ffill
    - 각 id에 대해 '비어있지 않은 첫 metadata dict'를 채택
    """
    if "id" not in df.columns or "metadata" not in df.columns:
        return {}

    tmp = df.copy()
    tmp["id"] = tmp["id"].ffill().astype(str)

    out: Dict[str, Dict[str, Any]] = {}
    for _, row in tmp.iterrows():
        _id = row["id"].strip()
        if not _id:
            continue
        meta_dict = _parse_metadata_cell(row["metadata"])
        if meta_dict and _id not in out:
            out[_id] = meta_dict
    return out


def _collect_note_by_id(df: pd.DataFrame) -> Dict[str, str]:
    """
    엑셀 DF에서 id별로 metadata 셀을 파싱해 note를 수집.
    - id는 ffill
    - 각 id에 대해 '비어있지 않은 첫 note'를 채택
    """
    if "id" not in df.columns or "metadata" not in df.columns:
        return {}

    tmp = df.copy()
    tmp["id"] = tmp["id"].ffill().astype(str)

    out: Dict[str, str] = {}
    for _, row in tmp.iterrows():
        _id = row["id"].strip()
        if not _id:
            continue
        meta_dict = _parse_metadata_cell(row["metadata"])
        note = str(meta_dict.get("note", "") or "").strip()
        if note and _id not in out:
            out[_id] = note
    return out


def _collect_medium_by_id(df: pd.DataFrame) -> Dict[str, str]:
    """
    엑셀에서 id별 Medium_category 값을 수집.
    - 병합 셀 보정을 위해 id, Medium_category ffill
    - 각 id에 대해 '비어있지 않은 첫 값'을 채택
    """
    if "id" not in df.columns:
        return {}

    tmp = df.copy()
    tmp["id"] = tmp["id"].ffill().astype(str)

    if "Medium_category" not in tmp.columns:
        return {}

    tmp["Medium_category"] = tmp["Medium_category"].ffill().fillna("").astype(str)

    out: Dict[str, str] = {}
    for _, row in tmp.iterrows():
        _id = row["id"].strip()
        mc = row["Medium_category"].strip()
        if _id and mc and _id not in out:
            out[_id] = mc
    return out


def _strip_brackets(s: str) -> str:
    if not s:
        return ""
    s = str(s).strip()
    if s.startswith("[") and s.endswith("]"):
        return s[1:-1].strip()
    return s


def _sort_label_keys(keys):
    # "설명 문장1", "설명 문장2" ... 같은 키를 숫자 기준 정렬
    def key_fn(k):
        k = str(k)
        m = re.search(r"(\d+)", k)
        return (0, int(m.group(1))) if m else (1, k)
    return sorted(list(keys), key=key_fn)


# =========================
# JSON -> Excel (정방향)
# =========================
def extract_sentences(doc: Dict[str, Any]) -> List[Tuple[str, str]]:
    """
    EX.exp_sentence 내부를 탐색해 (type, sentence) 튜플 리스트로 반환

    지원 형식:
    A) 구형: "[Type] sentence" 또는 "sentence"
    B) 신형:
       "exp_sentence": {
         "설명 문장1": {"feature": "[대상 식별 문장]", "sent": "..."},
         ...
       }
    """
    out: List[Tuple[str, str]] = []

    for ex in doc.get("EX", []):
        exp = ex.get("exp_sentence")
        if exp is None:
            continue

        # ===== 신형(dict: label -> {feature, sent}) =====
        if isinstance(exp, dict):
            is_new = any(isinstance(v, dict) and ("sent" in v or "feature" in v) for v in exp.values())
            if is_new:
                for label in _sort_label_keys(exp.keys()):
                    v = exp.get(label, {})
                    if not isinstance(v, dict):
                        continue
                    typ = _strip_brackets(v.get("feature", "") or "")
                    sent = str(v.get("sent", "") or "").strip()
                    if sent:
                        out.append((typ, sent))
                continue

        # ===== 구형(list[dict{...}] / dict / str) =====
        if isinstance(exp, list):
            for item in exp:
                if not isinstance(item, dict):
                    continue
                for _k, v in item.items():
                    seq = v if isinstance(v, list) else [v]
                    for s in seq:
                        if not s:
                            continue
                        text = str(s).strip()
                        m = TYPE_BRACKET_RE.match(text)
                        if m:
                            out.append((m.group(1).strip(), m.group(2).strip()))
                        else:
                            out.append(("", text))
            continue

        if isinstance(exp, dict):
            for k, v in exp.items():
                seq = v if isinstance(v, list) else [v]
                for s in seq:
                    if not s:
                        continue
                    text = str(s).strip()
                    m = TYPE_BRACKET_RE.match(text)
                    if m:
                        out.append((m.group(1).strip(), m.group(2).strip()))
                    else:
                        out.append(("", text))
            continue

        if isinstance(exp, str):
            text = exp.strip()
            if text:
                m = TYPE_BRACKET_RE.match(text)
                if m:
                    out.append((m.group(1).strip(), m.group(2).strip()))
                else:
                    out.append(("", text))

    return out


def _clean_url(u: str) -> str:
    if not u:
        return ""
    u = str(u).strip()
    if (u.startswith('"') and u.endswith('"')) or (u.startswith("'") and u.endswith("'")):
        u = u[1:-1].strip()
    return u


def format_metadata_and_url(meta: Dict[str, Any]) -> Tuple[str, str]:
    """
    metadata를 멀티라인 문자열로 정리하고, url만 분리해서 반환
    """
    url_only = _clean_url(meta.get("url", ""))
    lines = ['metadata : {']
    for k in META_ORDER:
        v = meta.get(k, "")
        if k == "url":
            v = url_only or meta.get("url", "")
        lines.append(f'  "{k}": "{v}",' if k != META_ORDER[-1] else f'  "{k}": "{v}"')
    lines.append("}")
    return "\n".join(lines), url_only


def extract_mdfcn_memo(mdfcn_infos):
    """
    mdfcn_infos[*].mdfcn_memo 가 JSON 문자열이면 value만 추출해
    "작업 목록 1 : ..." 형태로 번호 매겨 반환.
    항목 간에 빈 줄 한 줄 추가.
    """
    if not mdfcn_infos:
        return ""
    out, idx = [], 1
    for info in mdfcn_infos:
        raw = info.get("mdfcn_memo", "")
        if not raw:
            continue
        try:
            arr = json.loads(raw)
            if isinstance(arr, list):
                for obj in arr:
                    val = str((obj or {}).get("value", "")).strip()
                    if val:
                        out.append(f"작업 목록 {idx} : {val}")
                        idx += 1
        except Exception:
            val = str(raw).strip()
            if val:
                out.append(f"작업 목록 {idx} : {val}")
                idx += 1
    return "\n\n".join(out)


def to_rows(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    JSON(dict) -> 행 리스트
    """
    rows: List[Dict[str, Any]] = []
    docs = data.get("document", [])
    if not isinstance(docs, list):
        return rows

    for doc in docs:
        img_id = str(doc.get("id", ""))
        meta = doc.get("metadata", {}) or {}
        medium_category = str(meta.get("Medium_category", "") or "")
        worker_id_cnst = str(doc.get("worker_id_cnst", "") or "")

        md_text, md_url = format_metadata_and_url(meta)
        memo_text = extract_mdfcn_memo(doc.get("mdfcn_infos", []) or [])

        pairs = extract_sentences(doc) or [("", "")]
        for typ, sent in pairs:
            rows.append({
                "id": img_id,
                "worker_id_cnst": worker_id_cnst,
                "Medium_category": medium_category,
                "유형": typ,
                "설명 문장": sent,
                "metadata": md_text,
                "meta_url": md_url,  # 하이퍼링크용(엑셀 열에는 포함 안 함)
                "mdfcn_memo(검수자 수정 이력)": memo_text,
            })
    return rows


def estimate_wrapped_lines(text: str, col_chars: int) -> int:
    if not text:
        return 1
    total = 0
    width = max(col_chars, 5)
    for para in str(text).split("\n"):
        total += max(1, math.ceil(len(para) / (width * 1.08)))
    return max(1, total)


def _write_excel_to_bytes(all_rows: List[Dict[str, Any]]) -> bytes:
    """
    행 리스트 -> Excel bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "result"

    headers = [
        "id", "worker_id_cnst", "Medium_category",
        "유형", "설명 문장", "metadata", "mdfcn_memo\n(검수자 수정 이력)"
    ]
    ws.append(headers)

    # 헤더 스타일
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        cell.fill = HEADER_FILL

    # 열 너비(문자폭 기준 추정)
    widths = {1: 12, 2: 16, 3: 14, 4: 16, 5: 80, 6: 60, 7: 50}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 그룹 시작/개수 추적
    start_row_by_group: Dict[Tuple[str], int] = {}
    count_by_group: Dict[Tuple[str], int] = {}

    current_row = 2
    for row in all_rows:
        ws.append([
            xls_safe(row.get("id", "")),
            xls_safe(row.get("worker_id_cnst", "")),
            xls_safe(row.get("Medium_category", "")),
            xls_safe(row.get("유형", "")),
            xls_safe(row.get("설명 문장", "")),
            xls_safe(row.get("metadata", "")),
            xls_safe(row.get("mdfcn_memo(검수자 수정 이력)", "")),
        ])
        for c in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=c).alignment = Alignment(
                vertical="top", wrap_text=(c in (5, 6, 7))
            )
            ws.cell(row=current_row, column=c).border = THIN_BORDER

        key = (row.get("id", ""),)
        if key not in start_row_by_group:
            start_row_by_group[key] = current_row
            count_by_group[key] = 0
        count_by_group[key] += 1
        current_row += 1

    # 병합: 같은 id 블록에서 [id, worker, Medium_category, metadata, mdfcn_memo] 병합
    merge_cols = [1, 2, 3, 6, 7]
    for key, start in start_row_by_group.items():
        cnt = count_by_group[key]
        if cnt > 1:
            end = start + cnt - 1
            for col in merge_cols:
                ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)
                ws.cell(row=start, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    # metadata 하이퍼링크(같은 id 첫 행만)
    first_url_by_id: Dict[str, str] = {}
    for _r, row in enumerate(all_rows, start=2):
        rid = row.get("id", "")
        if rid and rid not in first_url_by_id:
            first_url_by_id[rid] = row.get("meta_url", "") or ""

    from openpyxl.cell.cell import MergedCell
    try:
        from openpyxl.worksheet.hyperlink import Hyperlink
    except Exception:
        Hyperlink = None  # 구버전 대비

    for key, start in start_row_by_group.items():
        doc_id = key[0]
        url = str(first_url_by_id.get(doc_id, "") or "").strip()
        if not (url and url.startswith(("http://", "https://"))):
            continue

        c = ws.cell(row=start, column=6)

        if isinstance(c, MergedCell):
            if Hyperlink is not None:
                ws._hyperlinks.append(Hyperlink(ref=c.coordinate, target=url, display=url))
        else:
            try:
                c.hyperlink = url
            except AttributeError:
                if Hyperlink is not None:
                    ws._hyperlinks.append(Hyperlink(ref=c.coordinate, target=url, display=url))

        c.font = Font(color=LINK_BLUE, underline="none")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = THIN_BORDER

    # 행 높이 대략 조정
    LINE_HEIGHT_PT = 18
    group_starts = set(start_row_by_group.values())
    for r in range(2, current_row):
        desc = ws.cell(row=r, column=5).value or ""
        desc_lines = estimate_wrapped_lines(desc, widths[5])
        if r in group_starts:
            meta_plain = ws.cell(row=r, column=6).value or ""
            memo_plain = ws.cell(row=r, column=7).value or ""
            need = max(
                desc_lines,
                estimate_wrapped_lines(meta_plain, widths[6]),
                estimate_wrapped_lines(memo_plain, widths[7]),
            )
        else:
            need = desc_lines
        ws.row_dimensions[r].height = max(1, need) * LINE_HEIGHT_PT

    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def photo_json_to_xlsx_bytes(data: Dict[str, Any]) -> bytes:
    """
    datalyManager에서 호출하는 공개 API
    """
    rows = to_rows(data)
    if not rows:
        return _write_excel_to_bytes([])
    return _write_excel_to_bytes(rows)


def _read_excel_multi(ef, sheet_name: Optional[Union[Iterable[str], str]] = None) -> pd.DataFrame:
    """
    Excel 파일에서 시트를 읽어 하나의 DataFrame으로 합친다.
    - sheet_name == None: 모든 시트
    - sheet_name == str: 해당 이름의 단일 시트
    - sheet_name == Iterable[str]: 지정 시트들만 순서대로
    반환 DF는 시트 순서와 원본 행 순서를 유지하도록 index를 다시 매깁니다.
    필요 컬럼(id, '설명 문장', 선택: '유형', 'Medium_category')이 없으면 빈 컬럼으로 보정.
    """
    need_cols = ["id", "설명 문장"]
    opt_cols = ["유형", "Medium_category"]

    if sheet_name is None:
        sheets = pd.read_excel(ef, sheet_name=None)
        dfs = []
        for name, df in sheets.items():
            df = df.copy()
            for c in need_cols + opt_cols:
                if c not in df.columns:
                    df[c] = ""
            df["__sheet__"] = str(name)
            dfs.append(df)
        if not dfs:
            return pd.DataFrame(columns=need_cols + opt_cols)
        return pd.concat(dfs, ignore_index=True)

    if isinstance(sheet_name, str):
        df = pd.read_excel(ef, sheet_name=sheet_name)
        df = df.copy()
        for c in need_cols + opt_cols:
            if c not in df.columns:
                df[c] = ""
        df["__sheet__"] = str(sheet_name)
        return df

    try:
        names = list(sheet_name)
    except TypeError:
        raise TypeError("sheet_name은 None, 문자열, 또는 문자열 리스트여야 합니다.")

    all_sheets = pd.read_excel(ef, sheet_name=None)
    dfs = []
    for nm in names:
        if nm not in all_sheets:
            continue
        df = all_sheets[nm].copy()
        for c in need_cols + opt_cols:
            if c not in df.columns:
                df[c] = ""
        df["__sheet__"] = str(nm)
        dfs.append(df)

    if not dfs:
        return pd.DataFrame(columns=need_cols + opt_cols)
    return pd.concat(dfs, ignore_index=True)


# ==========================================
# Excel('설명 문장') → JSON (역방향, ZIP 지원)
# ==========================================

def _delete_slot(slot_descriptor):
    """
    ('list', list_obj, idx)  -> list_obj.pop(idx)
    ('dict_scalar', dict_obj, key) -> dict_obj.pop(key, None)
    ('new_obj', exp_dict, label_key) -> exp_dict.pop(label_key, None)
    """
    mode = slot_descriptor[0]
    if mode == "list":
        lst, idx = slot_descriptor[1], slot_descriptor[2]
        if 0 <= idx < len(lst):
            lst.pop(idx)
    elif mode == "dict_scalar":
        obj, key = slot_descriptor[1], slot_descriptor[2]
        if isinstance(obj, dict):
            obj.pop(key, None)
    elif mode == "new_obj":
        container, label = slot_descriptor[1], slot_descriptor[2]
        if isinstance(container, dict):
            container.pop(label, None)


def _cleanup_exp_sentences(doc: Dict[str, Any]) -> None:
    """
    빈 문자열/빈 리스트/빈 딕셔너리를 걷어내서 exp_sentence 구조를 가볍게 정리.
    (키 자체도 제거)
    ✅ 신형(dict: label -> {feature,sent})도 정리
    """
    ex_list = doc.get("EX", [])
    if not isinstance(ex_list, list):
        return

    for ex in ex_list:
        exp = ex.get("exp_sentence")
        if exp is None:
            continue

        # 신형: dict(label -> {feature, sent})
        if isinstance(exp, dict):
            is_new = any(isinstance(v, dict) and ("sent" in v or "feature" in v) for v in exp.values())
            if is_new:
                new_exp = {}
                for k, v in exp.items():
                    if not isinstance(v, dict):
                        continue
                    feature = str(v.get("feature", "") or "").strip()
                    sent = str(v.get("sent", "") or "").strip()
                    # 둘 다 비면 제거
                    if not feature and not sent:
                        continue
                    new_exp[k] = {"feature": feature, "sent": sent}
                if new_exp:
                    ex["exp_sentence"] = new_exp
                else:
                    ex.pop("exp_sentence", None)
                continue

        # list 형태 (기존)
        if isinstance(exp, list):
            new_exp = []
            for item in exp:
                if isinstance(item, dict):
                    new_item = {}
                    for k, v in item.items():
                        if isinstance(v, list):
                            vv = [str(s).strip() for s in v if str(s or "").strip()]
                            if vv:
                                new_item[k] = vv
                        else:
                            s = str(v or "").strip()
                            if s:
                                new_item[k] = s
                    if new_item:
                        new_exp.append(new_item)
            ex["exp_sentence"] = new_exp
            if not new_exp:
                ex.pop("exp_sentence", None)

        # dict 형태 (구형 dict: key -> str/list[str])
        elif isinstance(exp, dict):
            new_exp = {}
            for k, v in exp.items():
                if isinstance(v, list):
                    vv = [str(s).strip() for s in v if str(s or "").strip()]
                    if vv:
                        new_exp[k] = vv
                else:
                    s = str(v or "").strip()
                    if s:
                        new_exp[k] = s
            if new_exp:
                ex["exp_sentence"] = new_exp
            else:
                ex.pop("exp_sentence", None)


def _compose_text_with_type(old_text: str, new_sentence: str, excel_type: str) -> str:
    """
    (구형 슬롯용) 엑셀 '유형'만 바꿔도 문장은 유지하면서 타입을 교체.
    """
    s_new = "" if new_sentence is None else str(new_sentence).strip()
    t_new = "" if excel_type is None else str(excel_type).strip()

    old_text_str = str(old_text or "")
    m = TYPE_BRACKET_RE.match(old_text_str)
    old_type = (m.group(1).strip() if m else "")
    old_body = (m.group(2).strip() if m else old_text_str.strip())

    if t_new.startswith("[") and t_new.endswith("]"):
        t_new = t_new[1:-1].strip()

    body = s_new if s_new else old_body
    final_type = t_new if t_new else old_type

    return f"[{final_type}] {body}".strip() if final_type else body


def _iter_sentence_slots_with_old(doc: Dict[str, Any]):
    """
    사진 JSON의 EX[*].exp_sentence에서 실제 '문장 슬롯'을 순서대로 찾아
    (slot_descriptor, old_text) 를 yield.

    ✅ 신형 슬롯:
      slot_descriptor = ("new_obj", exp_dict, label_key)
      old_text = (old_feature, old_sent)  # tuple
    """
    ex_list = doc.get("EX", [])
    if not isinstance(ex_list, list):
        return

    for ex in ex_list:
        exp = ex.get("exp_sentence")
        if exp is None:
            continue

        # ✅ 신형: dict(label -> {feature, sent})
        if isinstance(exp, dict):
            is_new = any(isinstance(v, dict) and ("sent" in v or "feature" in v) for v in exp.values())
            if is_new:
                for label in _sort_label_keys(exp.keys()):
                    obj = exp.get(label, {})
                    if not isinstance(obj, dict):
                        continue
                    old_feature = "" if obj.get("feature") is None else str(obj.get("feature"))
                    old_sent = "" if obj.get("sent") is None else str(obj.get("sent"))
                    yield (("new_obj", exp, label), (old_feature, old_sent))
                continue

        # 구형: list[ dict{key: list[str] or str}, ... ]
        if isinstance(exp, list):
            for item in exp:
                if isinstance(item, dict):
                    for k, v in item.items():
                        if isinstance(v, list):
                            for i, s in enumerate(v):
                                yield (("list", v, i), ("" if s is None else str(s)))
                        else:
                            yield (("dict_scalar", item, k), ("" if v is None else str(v)))
            continue

        # 구형: dict(key -> list[str]/str)
        elif isinstance(exp, dict):
            for k, v in exp.items():
                if isinstance(v, list):
                    for i, s in enumerate(v):
                        yield (("list", v, i), ("" if s is None else str(s)))
                else:
                    yield (("dict_scalar", exp, k), ("" if v is None else str(v)))

        elif isinstance(exp, str):
            yield (("dict_scalar", ex, "exp_sentence"), exp)


def _assign_text_to_slot(slot_descriptor, new_text: str):
    mode = slot_descriptor[0]
    if mode == "list":
        lst, idx = slot_descriptor[1], slot_descriptor[2]
        lst[idx] = new_text
    elif mode == "dict_scalar":
        obj, key = slot_descriptor[1], slot_descriptor[2]
        obj[key] = new_text
    # "new_obj"는 apply_excel_desc_to_photo_json에서 직접 처리


def _collect_excel_pairs_by_id(df: pd.DataFrame, skip_blank: bool = True) -> Dict[str, List[Tuple[str, str]]]:
    """
    엑셀에서 id별 (유형, 설명 문장) 시퀀스를 원본 행 순서대로 수집.
    - 병합 셀로 인해 비는 id/유형은 ffill로 채움
    - skip_blank=True면 빈 '설명 문장'은 건너뜀
    반환: { id: [(type, sentence), ...], ... }
    """
    required = {"id", "설명 문장"}
    if not required.issubset(set(df.columns)):
        raise ValueError("엑셀에 'id', '설명 문장' 컬럼이 필요합니다.")

    tmp = df.copy()
    if "id" in tmp.columns:
        tmp["id"] = tmp["id"].ffill()
    if "유형" in tmp.columns:
        tmp["유형"] = tmp["유형"].ffill()

    tmp["id"] = tmp["id"].astype(str)
    if "유형" not in tmp.columns:
        tmp["유형"] = ""
    tmp["유형"] = tmp["유형"].fillna("").astype(str)
    tmp["설명 문장"] = tmp["설명 문장"].fillna("").astype(str)

    bucket: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
    for _, row in tmp.iterrows():
        _id = row["id"].strip()
        typ = row["유형"].strip()
        sent = row["설명 문장"].strip()
        if skip_blank and not sent:
            continue
        bucket[_id].append((typ, sent))
    return bucket


def apply_excel_desc_to_photo_json(
    json_obj: Dict[str, Any],
    excel_df: pd.DataFrame,
    skip_blank: bool = False
) -> Dict[str, Any]:
    """
    사진 JSON에 엑셀의 '설명 문장'과 'Medium_category'를 반영.
    - 같은 id 내에서 '엑셀 행 순서'와 '기존 JSON 슬롯 순서'를 1:1로 맞춰 반영

    ✅ 신형 exp_sentence도 지원:
      exp_sentence[label] = {"feature": "...", "sent": "..."}
      - feature는 엑셀 '유형'을 "[...]"로 감싸 저장(엑셀에 대괄호가 없어도 자동)
      - sent는 엑셀 '설명 문장' 그대로 저장
    """
    mapping = _collect_excel_pairs_by_id(excel_df, skip_blank=skip_blank)
    metadata_map = _collect_metadata_by_id(excel_df)
    medium_map = _collect_medium_by_id(excel_df)
    note_map = _collect_note_by_id(excel_df)

    docs = json_obj.get("document", [])
    if not isinstance(docs, list):
        return json_obj

    for doc in docs:
        doc_id = str(doc.get("id", ""))

        # metadata dict 보장
        if doc_id and (doc_id in metadata_map or doc_id in medium_map or doc_id in note_map or not skip_blank):
            if not isinstance(doc.get("metadata"), dict):
                doc["metadata"] = {}
            meta_obj = doc["metadata"]
        else:
            meta_obj = None

        # 2-1) 엑셀 metadata 전체 반영
        if meta_obj is not None:
            meta_from_excel = metadata_map.get(doc_id)
            if meta_from_excel:
                meta_obj.update(meta_from_excel)

            mc_val = medium_map.get(doc_id, "")
            if mc_val or not skip_blank:
                meta_obj["Medium_category"] = mc_val

            note_val = note_map.get(doc_id, "")
            if note_val or not skip_blank:
                meta_obj["note"] = note_val

        # 2-2) 설명 문장/유형 반영
        seq = mapping.get(doc_id, [])
        slots = list(_iter_sentence_slots_with_old(doc))

        # exp_sentence가 전혀 없고, 엑셀 시퀀스가 있으면 신형 구조로 생성
        if not slots and seq:
            ex_list = doc.get("EX")
            if not isinstance(ex_list, list) or not ex_list:
                doc["EX"] = [{"exp_sentence": {}}]
                ex_list = doc["EX"]

            ex0 = ex_list[0]
            if "exp_sentence" not in ex0 or ex0["exp_sentence"] is None:
                ex0["exp_sentence"] = {}

            # 신형 dict로 강제
            exp = ex0["exp_sentence"]
            if not isinstance(exp, dict):
                exp = {}
                ex0["exp_sentence"] = exp

            # 설명 문장1..n 생성
            idx = 1
            for typ, sent in seq:
                typ = (typ or "").strip()
                sent = (sent or "").strip()
                if not (typ or sent):
                    continue

                t = typ
                if t.startswith("[") and t.endswith("]"):
                    t = t[1:-1].strip()
                feature_out = f"[{t}]" if t else ""

                exp[f"설명 문장{idx}"] = {"feature": feature_out, "sent": sent}
                idx += 1

            _cleanup_exp_sentences(doc)
            continue

        n = min(len(seq), len(slots))
        delete_slot_indices = []

        for i in range(n):
            (slot_desc, old_val) = slots[i]
            typ, new_sent = seq[i]
            typ_clean = (typ or "").strip()
            sent_clean = (new_sent or "").strip()

            # ✅ 신형 슬롯 처리
            if slot_desc[0] == "new_obj":
                exp_dict, label = slot_desc[1], slot_desc[2]
                old_feature, old_sent = old_val

                if typ_clean == "" and sent_clean == "":
                    delete_slot_indices.append(i)
                    continue

                # feature: 엑셀 유형 있으면 교체, 없으면 기존 유지
                if typ_clean:
                    t = typ_clean
                    if t.startswith("[") and t.endswith("]"):
                        t = t[1:-1].strip()
                    feature_out = f"[{t}]" if t else ""
                else:
                    feature_out = str(old_feature or "").strip()

                # sent: 엑셀 문장 있으면 교체, 없으면 기존 유지
                sent_out = sent_clean if sent_clean else str(old_sent or "").strip()

                obj = exp_dict.get(label)
                if not isinstance(obj, dict):
                    obj = {}
                    exp_dict[label] = obj
                obj["feature"] = feature_out
                obj["sent"] = sent_out
                continue

            # ===== 구형 슬롯 처리(문자열) =====
            (slot_desc2, old_text) = slots[i]
            if typ_clean == "" and sent_clean == "":
                delete_slot_indices.append(i)
                continue

            composed = _compose_text_with_type(old_text, sent_clean, typ_clean)
            _assign_text_to_slot(slot_desc2, composed)

        # 엑셀 행 수 < JSON 슬롯 수면, 남은 슬롯은 뒤에서부터 삭제해 개수 일치
        if len(slots) > len(seq):
            delete_slot_indices.extend(range(n, len(slots)))

        for idx in sorted(delete_slot_indices, reverse=True):
            _delete_slot(slots[idx][0])

        _cleanup_exp_sentences(doc)

    return json_obj


def apply_excel_desc_to_json_from_zip(
    zip_bytes: bytes,
    sheet_name: Optional[str] = None,
    skip_blank: bool = True,
) -> Tuple[bytes, str]:
    """
    (사진 전용) ZIP(엑셀+단일 JSON)을 받아 엑셀의 '설명 문장'을 JSON에 반영.
    반환: (updated_json_bytes, suggested_filename)
    """
    if not isinstance(zip_bytes, (bytes, bytearray)):
        raise TypeError("zip_bytes는 bytes/bytearray여야 합니다.")

    with zipfile.ZipFile(BytesIO(zip_bytes), "r") as zf:
        json_members = [m for m in zf.namelist() if m.lower().endswith(".json")]
        xlsx_members = [m for m in zf.namelist() if m.lower().endswith(".xlsx")]
        xls_members = [m for m in zf.namelist() if m.lower().endswith(".xls")]

        json_member = None
        for m in json_members:
            if Path(m).name.startswith("project_"):
                json_member = m
                break
        if json_member is None and json_members:
            json_member = json_members[0]

        excel_member = xlsx_members[0] if xlsx_members else (xls_members[0] if xls_members else None)

        if not json_member:
            raise FileNotFoundError("ZIP 안에 JSON 파일이 없습니다.")
        if not excel_member:
            raise FileNotFoundError("ZIP 안에 Excel 파일(.xlsx/.xls)이 없습니다.")

        with zf.open(json_member) as jf:
            json_obj = json.loads(jf.read().decode("utf-8"))

        with zf.open(excel_member) as ef:
            df = _read_excel_multi(ef, sheet_name=sheet_name)

        updated = apply_excel_desc_to_photo_json(json_obj, df, skip_blank=skip_blank)

        base = Path(json_member).name
        out_name = (base[:-5] if base.lower().endswith(".json") else base) + "_updated.json"

        text = json.dumps(updated, ensure_ascii=False, indent=2)
        return text.encode("utf-8"), out_name
