#dataly_tools/newspaper_eval_merged.py
import os
import json
import glob
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side

def get_team_and_worker(folder_name):
    return folder_name[0], folder_name[1:]

# 팀별 배경색 정의
TEAM_FILLS = {
    'A': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),  # 옅은 회색
    'B': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),  # 옅은 남색
    'C': PatternFill(start_color="FFE4C4", end_color="FFE4C4", fill_type="solid"),  # 옅은 주황색
}

# 테두리 스타일: 얇은 선
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def write_eval_table(ws, row_start, team_label, data, center, left):
    team = team_label[0]
    fill_color = TEAM_FILLS.get(team, PatternFill(fill_type=None))

    headers = ["순번", "평가준거", "평가항목", "점수 (1~7)", "근거", "점수 (1~7)", "근거"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_start, column=col, value=header)
        cell.alignment = center
        cell.fill = fill_color
        cell.border = thin_border

    sc1 = data.get("SC1", {})
    sc2 = data.get("SC2", {})

    sc1_ai_flag = sc1.get("ai_flag", False)
    sc2_ai_flag = sc2.get("ai_flag", False)

    sc1_eval = sc1.get("evaluation", {})
    sc2_eval = sc2.get("evaluation", {})

    cat_map = {"content": "내용", "organization": "조직", "expression": "표현"}

    criterion_map = {
        "description": "문제상황",
        "claims": "주장",
        "arguments": "논거 또는 실천 방안",
        "completion": "긴밀성 및 완결성",
        "accuracy": "문장 및 어휘"
    }

    current_row = row_start + 1
    score1_rows = []
    score2_rows = []

    for cat_eng, cat_kor in cat_map.items():
        sc1_cat = sc1_eval.get(cat_eng, {})
        sc2_cat = sc2_eval.get(cat_eng, {})

        if cat_eng == "content":
            comments_sc1 = sc1_cat.get("comments", {})
            comments_sc2 = sc2_cat.get("comments", {})
        else:
            comments_sc1 = {}
            comments_sc2 = {}

        score_keys = [k for k in sc1_cat.keys() if k not in ("comment", "comments")]
        if not score_keys:
            continue

        cat_start_row = current_row

        for idx, criterion_eng in enumerate(score_keys, 1):
            score1 = sc1_cat.get(criterion_eng)
            score2 = sc2_cat.get(criterion_eng)
            criterion_kor = criterion_map.get(criterion_eng, criterion_eng)

            if cat_eng == "content":
                sc1_comment = comments_sc1.get(criterion_eng, "")
                sc2_comment = comments_sc2.get(criterion_eng, "")
            else:
                sc1_comment = sc1_cat.get("comment", "")
                sc2_comment = sc2_cat.get("comment", "")

            row_num = current_row
            score1_rows.append(row_num)
            score2_rows.append(row_num)

            cell = ws.cell(row=row_num, column=2, value=str(idx))
            cell.alignment = center
            cell.fill = fill_color
            cell.border = thin_border

            cell = ws.cell(row=row_num, column=3, value=criterion_kor)
            cell.alignment = center
            cell.fill = fill_color
            cell.border = thin_border

            cell = ws.cell(row=row_num, column=4, value=score1 if isinstance(score1, (int, float)) else None)
            cell.alignment = center
            cell.fill = fill_color
            cell.border = thin_border

            cell = ws.cell(row=row_num, column=5, value=sc1_comment)
            cell.alignment = left
            cell.fill = fill_color
            cell.border = thin_border

            cell = ws.cell(row=row_num, column=6, value=score2 if isinstance(score2, (int, float)) else None)
            cell.alignment = center
            cell.fill = fill_color
            cell.border = thin_border

            cell = ws.cell(row=row_num, column=7, value=sc2_comment)
            cell.alignment = left
            cell.fill = fill_color
            cell.border = thin_border

            current_row += 1

        # 카테고리명 병합
        ws.merge_cells(start_row=cat_start_row, start_column=2, end_row=current_row - 1, end_column=2)
        cell = ws.cell(row=cat_start_row, column=2, value=cat_kor)
        cell.alignment = center
        cell.fill = fill_color
        cell.border = thin_border

        # 내용 총점 행 (엑셀 수식으로 계산)
        if cat_eng == "content":
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
            cell = ws.cell(row=current_row, column=2, value=f"{cat_kor} 총점")
            cell.alignment = center
            cell.fill = PatternFill(fill_type=None)
            cell.border = thin_border

            # 점수 합계 수식
            d_cells = [f"D{r}" for r in range(cat_start_row, current_row)]
            f_cells = [f"F{r}" for r in range(cat_start_row, current_row)]

            cell = ws.cell(row=current_row, column=4, value=f"=SUM({','.join(d_cells)})")
            cell.alignment = center
            cell.border = thin_border

            cell = ws.cell(row=current_row, column=6, value=f"=SUM({','.join(f_cells)})")
            cell.alignment = center
            cell.border = thin_border

            # 빈칸
            for col in [5, 7]:
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border

            current_row += 1

    # 전체 총점 행 (엑셀 수식으로 계산)
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    cell = ws.cell(row=current_row, column=2, value="전체 총점")
    cell.alignment = center
    cell.fill = fill_color
    cell.border = thin_border

    d_sum = "+".join([f"D{r}" for r in score1_rows])
    f_sum = "+".join([f"F{r}" for r in score2_rows])

    cell = ws.cell(row=current_row, column=4, value=f"=ROUND(({d_sum}-5)*16.7/5, 1)")
    cell.number_format = '0.0'
    cell.alignment = center
    cell.fill = fill_color
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=6, value=f"=ROUND(({f_sum}-5)*16.7/5, 1)")
    cell.number_format = '0.0'
    cell.alignment = center
    cell.fill = fill_color
    cell.border = thin_border

    for col in [5, 7]:
        cell = ws.cell(row=current_row, column=col)
        cell.fill = fill_color
        cell.border = thin_border

    # 팀 라벨 병합
    total_rows = current_row - (row_start + 1) + 1
    if total_rows > 0:
        ws.merge_cells(start_row=row_start + 1, start_column=1, end_row=current_row, end_column=1)
        cell = ws.cell(row=row_start + 1, column=1, value=team_label)
    else:
        cell = ws.cell(row=current_row, column=1, value=team_label)
        current_row += 1

    cell.alignment = center
    cell.fill = fill_color
    cell.border = thin_border

    return current_row + 1, current_row



def json_to_excel_stacked(root_path, week_num, storage_folder):
    def get_team_and_worker(folder_name):
        return folder_name[0], folder_name[1:]

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    all_folders = [f for f in os.listdir(root_path) if os.path.isdir(os.path.join(root_path, f))]
    workers = {}
    for folder in all_folders:
        if len(folder) >= 4:
            team, worker_id = get_team_and_worker(folder)
            workers.setdefault(worker_id, {})[team] = folder

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)  # 기본 시트 제거

    incomplete_records = []  # 미완료 문서 기록 리스트

    # 작업자별 정렬
    for worker_id in sorted(workers, key=lambda x: int(''.join(filter(str.isdigit, x)))):
        teams = workers[worker_id]

        # 3개 팀 폴더에서 week 경로의 JSON 파일명 통합 수집
        all_json_files = set()
        for team in ['A', 'B', 'C']:
            folder = teams.get(team)
            if not folder:
                continue
            json_dir = os.path.join(root_path, folder, f"week{week_num:02d}_{folder}", storage_folder)
            if not os.path.exists(json_dir):
                continue
            files = glob.glob(os.path.join(json_dir, "*.json"))
            files += glob.glob(os.path.join(json_dir, "storageX", "*.json"))
            files = [os.path.basename(f) for f in files]
            all_json_files.update(files)

        if not all_json_files:
            continue  # 세 팀 모두 문서 없으면 스킵

        sheet_name = "W" + (worker_id if not worker_id[0].isalpha() else worker_id[1:])
        ws = wb.create_sheet(title=sheet_name)

        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='top', wrap_text=True)

        start_row = 1
        label_index = 1

        for base_fname in sorted(all_json_files):
            team_data = {}

            # 각 팀별 json 불러오기 시도
            for team in ['A', 'B', 'C']:
                folder = teams.get(team)
                if not folder:
                    continue
                json_dir = os.path.join(root_path, folder, f"week{week_num:02d}_{folder}", storage_folder)
                json_path = os.path.join(json_dir, base_fname)
                json_path_storageX = os.path.join(json_dir, "storageX", base_fname)

                if os.path.exists(json_path):
                    with open(json_path, encoding='utf-8') as f:
                        team_data[team] = json.load(f)
                    team_data[team]['_incomplete'] = False
                elif os.path.exists(json_path_storageX):
                    with open(json_path_storageX, encoding='utf-8') as f:
                        team_data[team] = json.load(f)
                    team_data[team]['_incomplete'] = True
                    doc_id = team_data[team].get("id", "")
                    incomplete_records.append({
                        "doc_id": doc_id,
                        "filename": base_fname,
                        "team": team,
                        "worker": worker_id,
                        "sc1_missing": not team_data[team].get("SC1", {}).get("evaluation"),
                        "sc2_missing": not team_data[team].get("SC2", {}).get("evaluation")
                    })

            if not team_data:
                continue  # 세 팀 모두 해당 문서 없음

            print(f"[{worker_id}] 사용된 파일명: {base_fname}")

            # 대표 doc_id, 제목, 본문은 가능한 A팀 기준으로, 없으면 B팀, 없으면 C팀 순으로
            doc_source = None
            for t in ['A','B','C']:
                if t in team_data:
                    doc_source = team_data[t]
                    break

            if not doc_source:
                continue

            doc_id = doc_source.get("id", "")

            paragraphs = doc_source.get("paragraph", [])
            if paragraphs:
                title = paragraphs[0].get("form", "")
                body = "".join(p.get("form", "") for p in paragraphs[1:])
            else:
                title = ""
                body = ""

            def get_summary(team_data, sc_key):
                for t in ['A', 'B', 'C']:
                    if t in team_data and sc_key in team_data[t]:
                        summary = team_data[t][sc_key].get("summary", "")
                        if summary:
                            return summary
                return ""

            sc1_summary = get_summary(team_data, "SC1")
            sc2_summary = get_summary(team_data, "SC2")

            metadata = doc_source.get("metadata", {})

            # 문서번호
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
            cell = ws.cell(row=start_row, column=1, value="문서번호")
            cell.alignment = center

            ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row, end_column=7)
            cell = ws.cell(row=start_row, column=4, value=doc_id)
            cell.alignment = center

            # 제목
            ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=3)
            cell = ws.cell(row=start_row + 1, column=1, value="제목")
            cell.alignment = center

            ws.merge_cells(start_row=start_row + 1, start_column=4, end_row=start_row + 1, end_column=7)
            cell = ws.cell(row=start_row + 1, column=4, value=title)
            cell.alignment = center

            # 신문 사설 정보 라벨 (H열)
            cell = ws.cell(row=start_row + 1, column=8, value="신문 사설 정보")
            cell.alignment = center

            # 신문 사설 정보 내용 (H열 아래)
            if metadata:
                info_lines = [f"{k}: {v}" for k, v in metadata.items()]
                info_text = "\n".join(info_lines)
                cell = ws.cell(row=start_row + 2, column=8, value=info_text)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                ws.row_dimensions[start_row + 2].height = 15 * len(info_lines)

            # 원문
            ws.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 2, end_column=3)
            cell = ws.cell(row=start_row + 2, column=1, value="원문")
            cell.alignment = center

            ws.merge_cells(start_row=start_row + 2, start_column=4, end_row=start_row + 2, end_column=7)
            cell = ws.cell(row=start_row + 2, column=4, value=body)
            cell.alignment = left
            ws.row_dimensions[start_row + 2].height = 140

            # 요약문 작성자
            ws.merge_cells(start_row=start_row + 3, start_column=1, end_row=start_row + 3, end_column=3)
            cell = ws.cell(row=start_row + 3, column=1, value="요약문 작성자")
            cell.alignment = center

            ws.merge_cells(start_row=start_row + 3, start_column=4, end_row=start_row + 3, end_column=5)
            cell = ws.cell(row=start_row + 3, column=4, value="A")
            cell.alignment = center

            ws.merge_cells(start_row=start_row + 3, start_column=6, end_row=start_row + 3, end_column=7)
            cell = ws.cell(row=start_row + 3, column=6, value="B")
            cell.alignment = center

            # 요약문
            ws.merge_cells(start_row=start_row + 4, start_column=1, end_row=start_row + 4, end_column=3)
            cell = ws.cell(row=start_row + 4, column=1, value="요약문")
            cell.alignment = center

            ws.merge_cells(start_row=start_row + 4, start_column=4, end_row=start_row + 4, end_column=5)
            cell = ws.cell(row=start_row + 4, column=4, value=sc1_summary)
            cell.alignment = left

            ws.merge_cells(start_row=start_row + 4, start_column=6, end_row=start_row + 4, end_column=7)
            cell = ws.cell(row=start_row + 4, column=6, value=sc2_summary)
            cell.alignment = left

            ws.row_dimensions[start_row + 4].height = 140

            # 요약문 글자수
            ws.merge_cells(start_row=start_row + 5, start_column=1, end_row=start_row + 5, end_column=3)
            cell = ws.cell(row=start_row + 5, column=1, value="요약문 글자수")
            cell.alignment = center

            ws.merge_cells(start_row=start_row + 5, start_column=4, end_row=start_row + 5, end_column=5)
            cell = ws.cell(row=start_row + 5, column=4, value=len(sc1_summary))
            cell.alignment = center

            ws.merge_cells(start_row=start_row + 5, start_column=6, end_row=start_row + 5, end_column=7)
            cell = ws.cell(row=start_row + 5, column=6, value=len(sc2_summary))
            cell.alignment = center

            start_row += 6  # 빈 행 없이 이어서 작성

            def get_ai_flag_teams(team_data, sc_key):
                flagged_teams = []
                for t in ['A', 'B', 'C']:
                    if t in team_data and sc_key in team_data[t]:
                        if team_data[t][sc_key].get("ai_flag", False):
                            flagged_teams.append(f"{t}팀")
                return ", ".join(flagged_teams)

            sc1_ai_flag_teams = get_ai_flag_teams(team_data, "SC1")
            sc2_ai_flag_teams = get_ai_flag_teams(team_data, "SC2")

            # 라벨
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
            cell = ws.cell(row=start_row, column=1, value="생성 AI 의심")
            cell.alignment = center

            # SC1 - O 여부
            ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row, end_column=5)
            cell = ws.cell(row=start_row, column=4, value="O" if sc1_ai_flag_teams else "")
            cell.alignment = center

            # SC2 - O 여부
            ws.merge_cells(start_row=start_row, start_column=6, end_row=start_row, end_column=7)
            cell = ws.cell(row=start_row, column=6, value="O" if sc2_ai_flag_teams else "")
            cell.alignment = center

            # SC1 팀명 (H열)
            cell = ws.cell(row=start_row, column=8, value=sc1_ai_flag_teams)
            cell.alignment = left

            # SC2 팀명 (I열)
            cell = ws.cell(row=start_row, column=9, value=sc2_ai_flag_teams)
            cell.alignment = left
            start_row += 1

            total_score_cells_D = []  # SC1 전체 총점 셀 주소 저장
            total_score_cells_F = []  # SC2 전체 총점 셀 주소 저장

            # 평가표 작성 함수는 별도로 정의했다고 가정 (원래 쓰던 write_eval_table 재사용)
            for team, label_base in zip(['A', 'B', 'C'], ['A', 'B', 'C']):
                if team in team_data:
                    label = label_base + str(label_index)
                    start_row, total_score_row = write_eval_table(ws, start_row, label, team_data[team], center, left)

                    total_score_cells_D.append(f"D{total_score_row}")
                    total_score_cells_F.append(f"F{total_score_row}")

                    if team_data[team].get('_incomplete', False):
                        cell = ws.cell(row=start_row - 1, column=8, value="미완료")
                        cell.alignment = center

            label_index += 1

            # 팀 평균 행 추가
            if total_score_cells_D and total_score_cells_F:
                avg_row = start_row
                avg_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                ws.merge_cells(start_row=avg_row, start_column=1, end_row=avg_row, end_column=3)
                cell = ws.cell(row=avg_row, column=1, value="평균(A,B,C)")
                cell.alignment = center
                cell.fill = avg_fill
                cell.border = thin_border

                cell = ws.cell(row=avg_row, column=4, value=f"=ROUND(AVERAGE({','.join(total_score_cells_D)}), 1)")
                cell.number_format = '0.0'
                cell.alignment = center
                cell.fill = avg_fill
                cell.border = thin_border

                cell = ws.cell(row=avg_row, column=6, value=f"=ROUND(AVERAGE({','.join(total_score_cells_F)}), 1)")
                cell.number_format = '0.0'
                cell.alignment = center
                cell.fill = avg_fill
                cell.border = thin_border

                for col in [5, 7]:
                    cell = ws.cell(row=avg_row, column=col)
                    cell.fill = avg_fill
                    cell.border = thin_border

                start_row += 1

        col_widths = {"A":10,"B":14,"C":16,"D":12,"E":40,"F":12,"G":40,"H":30}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

    # --- 미완료 문서 시트 추가 ---
    if incomplete_records:
        ws_incomplete = wb.create_sheet(title="미완료 문서")
        headers = ["문서번호", "파일명", "팀", "작업자 ID"]
        for col, h in enumerate(headers, 1):
            cell = ws_incomplete.cell(row=1, column=col, value=h)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row_idx, record in enumerate(incomplete_records, start=2):
            ws_incomplete.cell(row=row_idx, column=1, value=record["doc_id"]).alignment = center
            ws_incomplete.cell(row=row_idx, column=2, value=record["filename"]).alignment = center
            ws_incomplete.cell(row=row_idx, column=3, value=record["team"]).alignment = center
            ws_incomplete.cell(row=row_idx, column=4, value=record["worker"]).alignment = center

        for col_letter in ['A', 'B', 'C', 'D']:
            ws_incomplete.column_dimensions[col_letter].width = 20

    save_path = os.path.join(root_path, "summary_eval_all.xlsx")
    wb.save(save_path)
    print(f"✅ 저장 완료: {save_path}")

#
# # 사용 예시
# root_path = '/Users/dataly/Desktop/NewspaperEval'
# week_num = 1
# storage_folder = "storage0"
#
# json_to_excel_stacked(root_path, week_num, storage_folder)


if __name__ == "__main__":
    root_path = input("root_path 입력: ").strip()
    week_num = int(input("week_num 입력 (예: 1): ").strip())
    storage_folder = input("storage_folder 입력 (예: storage0): ").strip()
    result_msg = json_to_excel_stacked(root_path, week_num, storage_folder)
    print(result_msg)
